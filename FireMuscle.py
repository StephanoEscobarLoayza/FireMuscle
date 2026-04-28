import streamlit as st
import pandas as pd
import json
import os
from datetime import datetime, date, timedelta
import math
from supabase import create_client, Client

# ─── PAGE CONFIG ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="🔥 FireMuscle · StephanoEl",
    page_icon="🔥",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─── SUPABASE CLIENT ───────────────────────────────────────────────────────────
@st.cache_resource
def get_supabase() -> Client:
    url  = st.secrets["SUPABASE_URL"]
    key  = st.secrets["SUPABASE_KEY"]
    return create_client(url, key)

supabase = get_supabase()

# ─── CONSTANTS ─────────────────────────────────────────────────────────────────
DIAS = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado"]

RUTINA_BASE = {
    "Día 1": {
        "titulo": "Espalda",
        "ejercicios": [
            {"musculo": "Abdomen",   "nombre": "Pre Entreno",              "series": 2, "reps": "20, 20, 7, 15, 15", "descanso": "30 seg"},
            {"musculo": "Espalda",   "nombre": "Superman",                 "series": 2, "reps": "1 min",             "descanso": "10 seg"},
            {"musculo": "Espalda",   "nombre": "Dumbbell Row",             "series": 2, "reps": "1 min",             "descanso": "10 seg"},
            {"musculo": "Espalda",   "nombre": "Prone Cobra",              "series": 2, "reps": "1 min",             "descanso": "10 seg"},
            {"musculo": "Espalda",   "nombre": "Dumbbell Pullover",        "series": 2, "reps": "1 min",             "descanso": "10 seg"},
            {"musculo": "Espalda",   "nombre": "Seal Flaps",               "series": 2, "reps": "1 min",             "descanso": "10 seg"},
            {"musculo": "Espalda",   "nombre": "Reverse Fly",              "series": 2, "reps": "1 min",             "descanso": "10 seg"},
            {"musculo": "Espalda",   "nombre": "Reverse Snow Angels",      "series": 2, "reps": "1 min",             "descanso": "10 seg"},
            {"musculo": "Espalda",   "nombre": "Remo",                     "series": 2, "reps": "1 min",             "descanso": "10 seg"},
            {"musculo": "Espalda",   "nombre": "Prone Thoracic Rotation",  "series": 2, "reps": "1 min",             "descanso": "10 seg"},
            {"musculo": "Espalda",   "nombre": "Peso muerto",              "series": 2, "reps": "1 min",             "descanso": "10 seg"},
            {"musculo": "Espalda",   "nombre": "Lat Pulldown",             "series": 6, "reps": "20, 16",            "descanso": "10 seg"},
        ]
    },
    "Día 2": {
        "titulo": "Abdominales",
        "ejercicios": [
            {"musculo": "Abdomen", "nombre": "Pre Entreno",          "series": 3, "reps": "20, 20, 7, 15, 20", "descanso": "30 seg"},
            {"musculo": "Abdomen", "nombre": "Abdomen Crunch lvl 1", "series": 2, "reps": "20, 20",            "descanso": "10 seg"},
            {"musculo": "Abdomen", "nombre": "Abdomen Crunch lvl 2", "series": 2, "reps": "20, 20",            "descanso": "10 seg"},
            {"musculo": "Abdomen", "nombre": "Abdomen Crunch lvl 3", "series": 2, "reps": "20, 20",            "descanso": "10 seg"},
            {"musculo": "Abdomen", "nombre": "Abdomen Crunch lvl 4", "series": 2, "reps": "20, 20",            "descanso": "15 seg"},
        ]
    },
    "Día 3": {
        "titulo": "Pierna",
        "ejercicios": [
            {"musculo": "Abdomen", "nombre": "Pre Entreno",           "series": 3, "reps": "20, 20, 10-9, 20, 20", "descanso": "30 seg"},
            {"musculo": "Pierna",  "nombre": "Goblet Squat",          "series": 2, "reps": "1 min, 1 min",         "descanso": "30 seg"},
            {"musculo": "Pierna",  "nombre": "Single‑Leg Calf Raise", "series": 2, "reps": "1 min, 1 min",         "descanso": "30 seg"},
            {"musculo": "Pierna",  "nombre": "Curtsy Lunge",          "series": 4, "reps": "1 min, 1 min",         "descanso": "30 seg"},
            {"musculo": "Pierna",  "nombre": "Dumbbell Sumo Squat",   "series": 2, "reps": "1 min, 1 min",         "descanso": "30 seg"},
            {"musculo": "Pierna",  "nombre": "Cossack Squat",         "series": 2, "reps": "1 min, 1 min",         "descanso": "30 seg"},
            {"musculo": "Pierna",  "nombre": "Hip Thrust",            "series": 2, "reps": "1 min, 1 min",         "descanso": "30 seg"},
            {"musculo": "Pierna",  "nombre": "Kneeling Squat",        "series": 2, "reps": "1 min, 1 min",         "descanso": "30 seg"},
        ]
    },
    "Día 4": {
        "titulo": "Pecho y Abdominales",
        "ejercicios": [
            {"musculo": "Abdomen", "nombre": "Pre Entreno",              "series": 2, "reps": "20, 20, 10-9, 20, 20", "descanso": "30 seg"},
            {"musculo": "Pecho",   "nombre": "Dumbbell Forward Press",   "series": 2, "reps": "1 min. 1 min",         "descanso": "30 seg"},
            {"musculo": "Pecho",   "nombre": "Triple Remo",              "series": 2, "reps": "1 min. 1 min",         "descanso": "30 seg"},
            {"musculo": "Pecho",   "nombre": "Dumbbell Front Raise",     "series": 2, "reps": "1 min. 1 min",         "descanso": "30 seg"},
            {"musculo": "Pecho",   "nombre": "Remo Cerrado",             "series": 2, "reps": "1 min. 1 min",         "descanso": "30 seg"},
            {"musculo": "Pecho",   "nombre": "Dumbbell Floor Press",     "series": 2, "reps": "1 min. 1 min",         "descanso": "30 seg"},
            {"musculo": "Pecho",   "nombre": "Dumbbell Pullover",        "series": 2, "reps": "1 min. 1 min",         "descanso": "30 seg"},
            {"musculo": "Pecho",   "nombre": "Dumbbell Floor Press",     "series": 4, "reps": "1 min. 1 min",         "descanso": "30 seg"},
            {"musculo": "Pecho",   "nombre": "Push-up",                  "series": 2, "reps": "1 min. 1 min",         "descanso": "30 seg"},
            {"musculo": "Pecho",   "nombre": "Maquina",                  "series": 3, "reps": "10, 10, Fallo",        "descanso": "20 seg"},
        ]
    },
    "Día 5": {
        "titulo": "Bíceps y Tríceps",
        "ejercicios": [
            {"musculo": "Abdomen",  "nombre": "Pre Entreno",                          "series": 2, "reps": "20, 20, 10-9, 20, 20", "descanso": "30 seg"},
            {"musculo": "Bíceps",   "nombre": "Curl de bíceps con rotación",          "series": 2, "reps": "1 min, 1 min",         "descanso": "30 seg"},
            {"musculo": "Triceps",  "nombre": "Extensión de tríceps",                 "series": 2, "reps": "1 min, 1 min",         "descanso": "30 seg"},
            {"musculo": "Bíceps",   "nombre": "Curl de bíceps con giro",              "series": 2, "reps": "1 min, 1 min",         "descanso": "30 seg"},
            {"musculo": "Triceps",  "nombre": "Patada de tríceps",                    "series": 2, "reps": "1 min, 1 min",         "descanso": "30 seg"},
            {"musculo": "Bíceps",   "nombre": "Curl martillo",                        "series": 2, "reps": "1 min, 1 min",         "descanso": "30 seg"},
            {"musculo": "Triceps",  "nombre": "Curl martillo inclinado",              "series": 2, "reps": "1 min, 1 min",         "descanso": "30 seg"},
            {"musculo": "Bíceps",   "nombre": "Curl lateral alterno",                 "series": 2, "reps": "1 min, 1 min",         "descanso": "30 seg"},
            {"musculo": "Triceps",  "nombre": "Martillo acostado hasta el pecho",     "series": 2, "reps": "1 min, 1 min",         "descanso": "30 seg"},
            {"musculo": "Triceps",  "nombre": "Extensión de tríceps alterna",         "series": 2, "reps": "1 min, 1 min",         "descanso": "30 seg"},
            {"musculo": "Triceps",  "nombre": "Remo con mancuerna sentado",           "series": 2, "reps": "1 min, 1 min",         "descanso": "30 seg"},
        ]
    },
    "Día 6": {
        "titulo": "Recuperación Activa - Semi Descanso",
        "ejercicios": [
            {"musculo": "Abdomen",       "nombre": "Pre Entreno",  "series": 2, "reps": "20, 20, 10-9, 20, 20",       "descanso": "10 seg"},
            {"musculo": "Pecho",         "nombre": "Planchas",     "series": 0, "reps": "-",                           "descanso": "-"},
            {"musculo": "Triceps",       "nombre": "Planchas",     "series": 0, "reps": "-",                           "descanso": "-"},
            {"musculo": "Pecho",         "nombre": "Maquina",      "series": 6, "reps": "12, 12, 12, 12, 15, Fallo",  "descanso": "1 min"},
            {"musculo": "Espalda/Brazos","nombre": "Dead Hang",    "series": 2, "reps": "5, 5",                        "descanso": "30 seg"},
        ]
    },
}

DIA_SEMANA_MAP = {0: "Día 1", 1: "Día 2", 2: "Día 3", 3: "Día 4", 4: "Día 5", 5: "Día 6"}

NUTRI_OBJETIVOS_EXCEL = {
    "calorias": 1800, "grasas": 55, "colesterol": 300,
    "sodio": 1750, "carbohidratos": 180, "proteinas": 130, "azucar": 25, "fibra": 25,
}

# ─── CSS ───────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Bebas+Neue&family=DM+Sans:wght@300;400;500;600&family=JetBrains+Mono:wght@400;600&display=swap');

:root {
    --bg:#0a0e17; --surface:#111827; --surface2:#1a2332;
    --accent:#00e5a0; --accent2:#ff6b35; --accent3:#4facfe; --accent4:#a78bfa;
    --text:#e8f0fe; --muted:#6b7a99;
    --good:#00e5a0; --warn:#ffd166; --bad:#ff4757; --border:#1e2d45;
}
html,body,[class*="css"]{font-family:'DM Sans',sans-serif;background-color:var(--bg);color:var(--text);}
#MainMenu,footer{visibility:hidden;}
.block-container{padding-top:1rem!important;}

.app-header{background:linear-gradient(135deg,#0a0e17 0%,#1a0d05 50%,#0d0a00 100%);border-bottom:1px solid rgba(255,107,53,.3);padding:1.2rem 2rem;margin:-1rem -1rem 2rem -1rem;display:flex;align-items:center;gap:1rem;position:relative;overflow:hidden;}
.app-title{font-family:'Bebas Neue',sans-serif;font-size:2.2rem;letter-spacing:3px;background:linear-gradient(135deg,#ff6b35 0%,#ff4500 50%,#ffd166 100%);-webkit-background-clip:text;-webkit-text-fill-color:transparent;margin:0;line-height:1;filter:drop-shadow(0 0 12px rgba(255,100,0,.4));}
.app-subtitle{font-size:.75rem;color:var(--muted);letter-spacing:2px;text-transform:uppercase;font-weight:500;}
.user-badge{margin-left:auto;background:var(--surface2);border:1px solid var(--border);border-radius:50px;padding:.4rem 1rem;font-size:.8rem;color:var(--accent);font-family:'JetBrains Mono',monospace;letter-spacing:1px;}

.card{background:var(--surface);border:1px solid var(--border);border-radius:16px;padding:1.5rem;margin-bottom:1rem;position:relative;overflow:hidden;}
.card::before{content:'';position:absolute;top:0;left:0;right:0;height:2px;background:linear-gradient(90deg,#ff6b35,#ffd166);}
.card-orange::before{background:linear-gradient(90deg,var(--accent2),var(--warn));}
.card-blue::before{background:linear-gradient(90deg,var(--accent3),#a78bfa);}
.card-purple::before{background:linear-gradient(90deg,#a78bfa,#ec4899);}
.card-title{font-family:'Bebas Neue',sans-serif;font-size:1.3rem;letter-spacing:2px;color:var(--accent);margin-bottom:1rem;display:flex;align-items:center;gap:.5rem;}
.card-orange .card-title{color:var(--accent2);}
.card-blue .card-title{color:var(--accent3);}
.card-purple .card-title{color:var(--accent4);}

.pill{display:inline-block;padding:.2rem .7rem;border-radius:50px;font-size:.7rem;font-weight:700;letter-spacing:1px;text-transform:uppercase;}
.pill-good{background:rgba(0,229,160,.15);color:var(--good);border:1px solid rgba(0,229,160,.3);}
.pill-warn{background:rgba(255,209,102,.15);color:var(--warn);border:1px solid rgba(255,209,102,.3);}
.pill-bad{background:rgba(255,71,87,.15);color:var(--bad);border:1px solid rgba(255,71,87,.3);}
.pill-info{background:rgba(79,172,254,.15);color:var(--accent3);border:1px solid rgba(79,172,254,.3);}
.pill-purple{background:rgba(167,139,250,.15);color:var(--accent4);border:1px solid rgba(167,139,250,.3);}

.prog-wrap{margin:.4rem 0;}
.prog-label{display:flex;justify-content:space-between;font-size:.72rem;color:var(--muted);margin-bottom:.25rem;}
.prog-bar{height:6px;background:var(--border);border-radius:99px;overflow:hidden;}
.prog-fill{height:100%;border-radius:99px;transition:width .4s ease;}
.prog-good{background:linear-gradient(90deg,#00e5a0,#00c07d);}
.prog-warn{background:linear-gradient(90deg,#ffd166,#ffb347);}
.prog-bad{background:linear-gradient(90deg,#ff4757,#ff1f3a);}

.ex-entry{background:var(--surface2);border:1px solid var(--border);border-radius:12px;padding:.8rem 1rem;margin:.4rem 0;display:flex;align-items:center;gap:1rem;}
.ex-icon{font-size:1.4rem;}
.ex-info{flex:1;}
.ex-name{font-weight:600;font-size:.9rem;}
.ex-detail{font-size:.75rem;color:var(--muted);margin-top:.1rem;}

.food-row{background:var(--surface2);border:1px solid var(--border);border-radius:10px;padding:.7rem 1rem;margin:.4rem 0;display:flex;justify-content:space-between;align-items:center;}
.food-name{font-weight:500;font-size:.9rem;}
.food-cal{font-family:'JetBrains Mono',monospace;color:var(--accent);font-size:.85rem;}

.stTextInput input,.stNumberInput input,.stSelectbox select{background:var(--surface2)!important;border:1px solid var(--border)!important;color:var(--text)!important;border-radius:10px!important;}
.stButton>button{background:linear-gradient(135deg,var(--accent),#00c07d)!important;color:var(--bg)!important;font-weight:700!important;border:none!important;border-radius:10px!important;letter-spacing:1px!important;transition:opacity .2s!important;}
.stButton>button:hover{opacity:.85!important;}
.stSidebar{background:var(--surface)!important;border-right:1px solid var(--border);}

.metric-box{background:var(--surface2);border:1px solid var(--border);border-radius:14px;padding:1rem 1.2rem;text-align:center;}
.metric-val{font-family:'Bebas Neue',sans-serif;font-size:2rem;color:var(--accent);line-height:1;}
.metric-label{font-size:.68rem;color:var(--muted);text-transform:uppercase;letter-spacing:1.5px;margin-top:.3rem;}

.alert-good{background:rgba(0,229,160,.08);border:1px solid rgba(0,229,160,.25);border-radius:10px;padding:.8rem 1rem;color:var(--good);font-size:.85rem;margin:.5rem 0;}
.alert-warn{background:rgba(255,209,102,.08);border:1px solid rgba(255,209,102,.25);border-radius:10px;padding:.8rem 1rem;color:var(--warn);font-size:.85rem;margin:.5rem 0;}
.alert-bad{background:rgba(255,71,87,.08);border:1px solid rgba(255,71,87,.25);border-radius:10px;padding:.8rem 1rem;color:var(--bad);font-size:.85rem;margin:.5rem 0;}
.alert-info{background:rgba(79,172,254,.08);border:1px solid rgba(79,172,254,.25);border-radius:10px;padding:.8rem 1rem;color:var(--accent3);font-size:.85rem;margin:.5rem 0;}

.section-sep{border:none;border-top:1px solid var(--border);margin:1.5rem 0;}

.db-week-header{font-family:'Bebas Neue',sans-serif;font-size:1.6rem;letter-spacing:3px;background:linear-gradient(90deg,var(--accent4),#ec4899);-webkit-background-clip:text;-webkit-text-fill-color:transparent;margin:1.5rem 0 .5rem;}
.db-ex-row{display:flex;gap:1rem;padding:.4rem .3rem;border-bottom:1px solid var(--border);font-size:.82rem;align-items:center;}
.db-ex-row:last-child{border-bottom:none;}
.db-musculo{color:var(--accent3);font-weight:600;min-width:80px;}
.db-nombre{flex:1;color:var(--text);}
.db-detail{color:var(--muted);font-family:'JetBrains Mono',monospace;font-size:.75rem;}

.routine-badge{display:inline-flex;align-items:center;gap:.3rem;background:rgba(0,229,160,.1);border:1px solid rgba(0,229,160,.3);border-radius:6px;padding:.15rem .6rem;font-size:.7rem;color:var(--accent);font-weight:600;letter-spacing:.5px;}

[data-testid="stForm"] small,
[data-testid="stForm"] .st-emotion-cache-1gulkj5,
[data-testid="stForm"] .eyeqlp53,
small.eyeqlp53,
div[data-testid="InputInstructions"],
[data-testid="InputInstructions"] {
    display: none !important;
}
[data-testid="stToolbar"],
[data-testid="stDecoration"],
header[data-testid="stHeader"] {
    display: none !important;
}

[data-testid="stToolbar"],
[data-testid="stDecoration"],
header[data-testid="stHeader"] {
    display: none !important;
}

/* AGREGÁ ESTO JUSTO ANTES DEL CIERRE */
button[data-testid="baseButton-secondary"] p {
    font-size: 1rem !important;
    line-height: 1 !important;
}

</style>
""", unsafe_allow_html=True)

# ─── SUPABASE HELPERS ──────────────────────────────────────────────────────────

def sb_login(username, password):
    res = supabase.table("fm_usuarios").select("*").eq("username", username).eq("password", password).execute()
    return len(res.data) > 0

def sb_register(username, password, email):
    try:
        supabase.table("fm_usuarios").insert({"username": username, "password": password, "email": email}).execute()
        return True, ""
    except Exception as e:
        return False, str(e)

def sb_user_exists(username):
    res = supabase.table("fm_usuarios").select("username").eq("username", username).execute()
    return len(res.data) > 0

def sb_get_perfil(username):
    res = supabase.table("fm_perfil").select("*").eq("username", username).execute()
    return res.data[0] if res.data else {}

def sb_save_perfil(username, perfil: dict):
    existing = supabase.table("fm_perfil").select("id").eq("username", username).execute()
    perfil["username"] = username
    perfil["updated_at"] = datetime.now().isoformat()
    if existing.data:
        supabase.table("fm_perfil").update(perfil).eq("username", username).execute()
    else:
        supabase.table("fm_perfil").insert(perfil).execute()

def sb_get_alimentos(username, fecha: str):
    res = supabase.table("fm_alimentos").select("*").eq("username", username).eq("fecha", fecha).order("created_at").execute()
    return res.data or []

def sb_add_alimento(username, fecha: str, food: dict):
    food["username"] = username
    food["fecha"]    = fecha
    supabase.table("fm_alimentos").insert(food).execute()

def sb_update_alimento(row_id, food: dict):
    supabase.table("fm_alimentos").update(food).eq("id", row_id).execute()

def sb_delete_alimento(row_id):
    supabase.table("fm_alimentos").delete().eq("id", row_id).execute()

def sb_get_ejercicios(username, dia_key: str):
    res = supabase.table("fm_ejercicios").select("*").eq("username", username).eq("dia_key", dia_key).order("created_at").execute()
    return res.data or []

def sb_add_ejercicio(username, dia_key: str, ex: dict):
    ex["username"] = username
    ex["dia_key"]  = dia_key
    supabase.table("fm_ejercicios").insert(ex).execute()

def sb_delete_ejercicio(row_id):
    supabase.table("fm_ejercicios").delete().eq("id", row_id).execute()

def sb_clear_ejercicios(username, dia_key: str):
    supabase.table("fm_ejercicios").delete().eq("username", username).eq("dia_key", dia_key).execute()

def sb_get_database(username):
    res = supabase.table("fm_database").select("*").eq("username", username).order("semana_num").order("dia_label").execute()
    return res.data or []

def sb_save_database(username, semana_num: int, dia_label: str, ejercicios, alimentos, notas: str):
    payload = {
        "username":    username,
        "semana_num":  semana_num,
        "dia_label":   dia_label,
        "fecha":       date.today().isoformat(),
        "ejercicios":  ejercicios,
        "alimentos":   alimentos,
        "notas":       notas,
        "guardado_en": datetime.now().strftime("%H:%M:%S"),
    }
    existing = supabase.table("fm_database").select("id").eq("username", username).eq("semana_num", semana_num).eq("dia_label", dia_label).execute()
    if existing.data:
        supabase.table("fm_database").update(payload).eq("id", existing.data[0]["id"]).execute()
    else:
        supabase.table("fm_database").insert(payload).execute()

def sb_delete_database_row(row_id):
    supabase.table("fm_database").delete().eq("id", row_id).execute()
    
def sb_get_rutina(username):
    res = supabase.table("fm_rutinas").select("*").eq("username", username).order("dia_key").execute()
    return res.data or []

def sb_save_rutina_dia(username, dia_key: str, titulo: str, ejercicios: list):
    existing = supabase.table("fm_rutinas").select("id").eq("username", username).eq("dia_key", dia_key).execute()
    payload = {
        "username":   username,
        "dia_key":    dia_key,
        "titulo":     titulo,
        "ejercicios": ejercicios,
    }
    if existing.data:
        supabase.table("fm_rutinas").update(payload).eq("id", existing.data[0]["id"]).execute()
    else:
        supabase.table("fm_rutinas").insert(payload).execute()

def sb_init_rutina_usuario(username):
    # Usuarios nuevos empiezan con rutina vacía
    # Solo StephanoEl carga la rutina base por defecto
    if username == "StephanoEl":
        existing = sb_get_rutina(username)
        if not existing:
            for dia_key, dia_data in RUTINA_BASE.items():
                sb_save_rutina_dia(username, dia_key, dia_data["titulo"], dia_data["ejercicios"])

def get_rutina_usuario(username):
    rows = sb_get_rutina(username)
    if not rows:
        sb_init_rutina_usuario(username)
        rows = sb_get_rutina(username)
    rutina = {}
    for row in rows:
        rutina[row["dia_key"]] = {
            "titulo":     row["titulo"],
            "ejercicios": row["ejercicios"] or [],
            "id":         row["id"],
        }
    return rutina    

def sb_get_next_semana_num(username):
    res = supabase.table("fm_database").select("semana_num").eq("username", username).execute()
    if not res.data:
        return 1
    nums = [r["semana_num"] for r in res.data]
    return max(nums) + 1

# ─── NUTRIENT HELPERS ─────────────────────────────────────────────────────────
def calc_limits(age, weight_kg, height_cm, sex="Masculino", activity="Moderado", objetivo="Mantener peso", use_excel=False):
    if use_excel:
        return NUTRI_OBJETIVOS_EXCEL.copy()
    if sex == "Masculino":
        bmr = 10 * weight_kg + 6.25 * height_cm - 5 * age + 5
    else:
        bmr = 10 * weight_kg + 6.25 * height_cm - 5 * age - 161
    activity_factors = {"Sedentario": 1.2, "Ligero": 1.375, "Moderado": 1.55, "Activo": 1.725, "Muy activo": 1.9}
    tdee = bmr * activity_factors.get(activity, 1.55)

    if objetivo == "Perder grasa":
        tdee      = tdee * 0.85
        protein_g = weight_kg * 2.0
        fat_g     = (tdee * 0.25) / 9
    elif objetivo == "Ganar músculo":
        tdee      = tdee * 1.10
        protein_g = weight_kg * 1.8
        fat_g     = (tdee * 0.28) / 9
    else:  # Mantener peso
        protein_g = weight_kg * 1.8
        fat_g     = (tdee * 0.30) / 9

    carbs_g = (tdee - protein_g * 4 - fat_g * 9) / 4
    return {
        "calorias": round(tdee), "proteinas": round(protein_g),
        "carbohidratos": round(carbs_g), "grasas": round(fat_g),
        "azucar": round(carbs_g * 0.10), "fibra": 30 if sex == "Masculino" else 25,
        "sodio": 2300, "colesterol": 300,
    }

def get_status(consumed, limit, nutrient):
    pct = (consumed / limit * 100) if limit > 0 else 0
    if nutrient == "fibra":
        if consumed >= limit:        return "✅ Excelente", "good", "Meta alcanzada"
        elif consumed >= limit*0.6:  return "⚠️ Moderado",  "warn", f"Faltan {limit-consumed:.0f}g"
        else:                        return "❌ Bajo",       "bad",  "Muy poca fibra"
    if pct <= 80:    return "✅ Bien",      "good", "Dentro del límite"
    elif pct <= 100: return "⚠️ Moderado",  "warn", "Cerca del límite"
    elif pct <= 130: return "⚠️ Excedido",  "warn", f"Superado en {pct-100:.0f}%"
    else:            return "❌ Exceso",    "bad",  f"¡Muy alto! ({pct:.0f}%)"

def overall_diet_status(foods, limits):
    if not foods: return None
    totals = sum_nutrients(foods)
    issues, warns = 0, 0
    for k in ["grasas", "colesterol", "sodio", "carbohidratos", "azucar"]:
        if k in limits and limits[k] > 0:
            pct = totals.get(k, 0) / limits[k] * 100
            if pct > 130:   issues += 1
            elif pct > 100: warns  += 1
    fibra_pct = totals.get("fibra", 0) / limits.get("fibra", 30) * 100
    if fibra_pct < 60: warns += 1
    if issues == 0 and warns == 0: return "good", "🌟 Excelente alimentación hoy"
    elif issues == 0:              return "warn", "⚡ Alimentación moderada — pequeños ajustes"
    else:                          return "bad",  f"⚠️ {issues} nutriente(s) muy excedido(s) hoy"

def sum_nutrients(foods):
    keys = ["calorias","proteinas","carbohidratos","grasas","azucar","fibra","sodio","colesterol"]
    total = {k: 0 for k in keys}
    for f in foods:
        for k in keys:
            total[k] += float(f.get(k, 0) or 0)
    return total

def get_musculo_from_exercise(ex):
    return ex.get("musculo") or "—"

def get_today_key():
    return date.today().isoformat()

# ─── SESSION STATE ─────────────────────────────────────────────────────────────
if "logged_in" not in st.session_state:
    st.session_state.logged_in = False
if "username" not in st.session_state:
    st.session_state.username = ""
if "selected_day" not in st.session_state:
    dow = date.today().weekday()
    st.session_state.selected_day = DIA_SEMANA_MAP.get(dow, "Día 1")
if "editing_food_idx" not in st.session_state:
    st.session_state.editing_food_idx = None
if "food_form_counter" not in st.session_state:
    st.session_state.food_form_counter = 0
if "perfil_cache" not in st.session_state:
    st.session_state.perfil_cache = {}

# ═══════════════════════════════════════════════════════════════════════════════
# LOGIN
# ═══════════════════════════════════════════════════════════════════════════════
if not st.session_state.logged_in:
    st.markdown("""
    <style>
    [data-testid="stAppViewContainer"] {
        background-image: url("https://png.pngtree.com/background/20230528/original/pngtree-gym-is-reflected-in-an-odd-light-picture-image_2773779.jpg");
        background-size: cover; background-position: center; background-attachment: fixed;
    }
    [data-testid="stAppViewContainer"]::before {
        content:'';position:fixed;inset:0;
        background:linear-gradient(135deg,rgba(5,8,15,.90) 0%,rgba(20,8,3,.88) 50%,rgba(5,8,15,.92) 100%);
        z-index:0;pointer-events:none;
    }
    .block-container{position:relative;z-index:1;}
    .login-card-anim{animation:fadeSlideUp 0.6s cubic-bezier(0.16,1,0.3,1) both;}
    @keyframes fadeSlideUp{from{opacity:0;transform:translateY(28px) scale(0.97);}to{opacity:1;transform:translateY(0) scale(1);}}
    .stTabs [data-baseweb="tab-list"]{background:rgba(255,255,255,.04)!important;border-radius:10px!important;border:1px solid rgba(255,107,53,.2)!important;padding:3px!important;}
    .stTabs [data-baseweb="tab"]{border-radius:8px!important;color:#6b7a99!important;font-weight:600!important;}
    .stTabs [aria-selected="true"]{background:rgba(255,107,53,.18)!important;color:#ff8c55!important;}
    </style>
    """, unsafe_allow_html=True)

    col1, col2, col3 = st.columns([1, 1.8, 1])
    with col2:
        st.markdown("<br><br>", unsafe_allow_html=True)
        st.markdown("""
        <div class="login-card-anim" style="text-align:center;padding:1rem 0 .5rem;">
            <div style="font-family:'Bebas Neue',sans-serif;font-size:3rem;letter-spacing:4px;
                        background:linear-gradient(135deg,#ff6b35,#ffd166);
                        -webkit-background-clip:text;-webkit-text-fill-color:transparent;
                        filter:drop-shadow(0 0 16px rgba(255,100,0,.5));">
                🔥 FIREMUSCLE
            </div>
            <div style="font-size:.75rem;color:#6b7a99;letter-spacing:3px;text-transform:uppercase;margin-top:.3rem;">
                Forja tu mejor versión
            </div>
        </div>
        """, unsafe_allow_html=True)

        login_tab, register_tab = st.tabs(["🔑 Iniciar sesión", "✨ Crear cuenta"])

        with login_tab:
            st.markdown("<br>", unsafe_allow_html=True)
            with st.form(key="login_form", clear_on_submit=False):
                username = st.text_input("👤 Usuario", placeholder="Nombre de usuario")
                password = st.text_input("🔑 Contraseña", type="password", placeholder="••••••••")
                st.markdown("<br>", unsafe_allow_html=True)
                submitted_login = st.form_submit_button("INICIAR SESIÓN", use_container_width=True)

            if submitted_login:
                if sb_login(username, password):
                    st.session_state.logged_in = True
                    st.session_state.username  = username
                    st.session_state.perfil_cache = sb_get_perfil(username)
                    st.rerun()
                else:
                    st.markdown('<div class="alert-bad">❌ Usuario o contraseña incorrectos</div>', unsafe_allow_html=True)

        with register_tab:
            st.markdown("<br>", unsafe_allow_html=True)
            with st.form(key="register_form", clear_on_submit=False):
                new_user  = st.text_input("👤 Nombre de usuario", placeholder="Ej: StephanoEl")
                new_email = st.text_input("📧 Email", placeholder="tu@correo.com")
                new_pass  = st.text_input("🔑 Contraseña", type="password", placeholder="Mínimo 6 caracteres")
                new_pass2 = st.text_input("🔒 Confirmar contraseña", type="password", placeholder="Repite tu contraseña")
                st.markdown("<br>", unsafe_allow_html=True)
                submitted_register = st.form_submit_button("CREAR CUENTA", use_container_width=True)

            if submitted_register:
                if not new_user.strip():
                    st.markdown('<div class="alert-bad">❌ Ingresa un nombre de usuario</div>', unsafe_allow_html=True)
                elif sb_user_exists(new_user):
                    st.markdown('<div class="alert-bad">❌ Ese usuario ya existe</div>', unsafe_allow_html=True)
                elif len(new_pass) < 6:
                    st.markdown('<div class="alert-warn">⚠️ La contraseña debe tener al menos 6 caracteres</div>', unsafe_allow_html=True)
                elif new_pass != new_pass2:
                    st.markdown('<div class="alert-bad">❌ Las contraseñas no coinciden</div>', unsafe_allow_html=True)
                else:
                    ok, err = sb_register(new_user, new_pass, new_email)
                    if ok:
                        st.markdown('<div class="alert-good">✅ ¡Cuenta creada! Ahora inicia sesión.</div>', unsafe_allow_html=True)
                    else:
                        st.markdown(f'<div class="alert-bad">❌ Error: {err}</div>', unsafe_allow_html=True)

        st.markdown("""<div style="text-align:center;margin-top:1.5rem;font-size:.65rem;color:#3a4560;letter-spacing:1px;">🔥 FIREMUSCLE v2.0 · StephanoEl</div>""", unsafe_allow_html=True)

    st.stop()

# ═══════════════════════════════════════════════════════════════════════════════
# APP PRINCIPAL
# ═══════════════════════════════════════════════════════════════════════════════
current_user = st.session_state.username
gp = st.session_state.perfil_cache

st.markdown(f"""
<div class="app-header">
    <div>
        <div class="app-title">🔥 FIREMUSCLE</div>
        <div class="app-subtitle">Forja tu mejor versión · Salud & Rendimiento · v2.0</div>
    </div>
    <div class="user-badge">👤 {current_user}</div>
</div>
""", unsafe_allow_html=True)

# ─── SIDEBAR ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("""<div style="font-family:'Bebas Neue',sans-serif;font-size:1.1rem;letter-spacing:3px;color:#00e5a0;margin-bottom:1rem;">⚙️ PERFIL PERSONAL</div>""", unsafe_allow_html=True)
    age      = st.number_input("🎂 Edad",        min_value=10,   max_value=100,   value=int(gp.get("age", 25)),        step=1)
    weight   = st.number_input("⚖️ Peso (kg)",   min_value=30.0, max_value=250.0, value=float(gp.get("weight", 70.0)), step=0.5)
    height   = st.number_input("📏 Altura (cm)", min_value=100.0, max_value=220.0, value=float(gp.get("height", 170.0)), step=0.5)
    sex      = st.selectbox("⚥ Sexo", ["Masculino","Femenino"], index=0 if gp.get("sex","Masculino")=="Masculino" else 1)
    activity = st.selectbox("🏃 Nivel de actividad",
                            ["Sedentario","Ligero","Moderado","Activo","Muy activo"],
                            index=["Sedentario","Ligero","Moderado","Activo","Muy activo"].index(gp.get("activity","Moderado")))
    objetivo = st.selectbox("🎯 Objetivo", ["Perder grasa","Mantener peso","Ganar músculo"],
                            index=["Perder grasa","Mantener peso","Ganar músculo"].index(gp.get("objetivo","Mantener peso")))

    if current_user == "StephanoEl":
        use_excel_limits = st.toggle("📊 Usar objetivos del Excel", value=bool(gp.get("use_excel_limits", True)))
    else:
        use_excel_limits = False
    if st.button("💾 Guardar Perfil", use_container_width=True):
        perfil_data = {
            "age": age, "weight": weight, "height": height,
            "sex": sex, "activity": activity,
            "objetivo": objetivo,  # ← agregá esto
            "use_excel_limits": use_excel_limits,
            "deficit": int(gp.get("deficit", 0)),
        }
        sb_save_perfil(current_user, perfil_data)
        st.session_state.perfil_cache = perfil_data
        gp = perfil_data
        st.success("✅ Perfil guardado")

    limits = calc_limits(age, weight, height, sex, activity, objetivo=objetivo, use_excel=use_excel_limits)
    
    st.markdown("<hr class='section-sep'>", unsafe_allow_html=True)
    src_label = "📊 Del Excel" if use_excel_limits else "🧮 Calculado"
    st.markdown(f"""<div style="font-size:.7rem;color:#6b7a99;text-transform:uppercase;letter-spacing:2px;margin-bottom:.8rem;">Tus límites diarios <span class='pill pill-info'>{src_label}</span></div>""", unsafe_allow_html=True)

    for k, v in {
        "🔥 Calorías":   f"{limits['calorias']} kcal",
        "💪 Proteínas":  f"{limits['proteinas']}g",
        "🌾 Carbos":     f"{limits['carbohidratos']}g",
        "🧈 Grasas":     f"{limits['grasas']}g",
        "🍬 Azúcar":     f"{limits['azucar']}g",
        "🌿 Fibra":      f"{limits['fibra']}g",
        "🧂 Sodio":      f"{limits['sodio']}mg",
        "🫀 Colesterol": f"{limits['colesterol']}mg",
    }.items():
        st.markdown(f"""<div style="display:flex;justify-content:space-between;padding:.35rem 0;border-bottom:1px solid #1e2d45;font-size:.8rem;"><span style="color:#6b7a99;">{k}</span><span style="font-family:'JetBrains Mono',monospace;color:#e8f0fe;font-size:.78rem;">{v}</span></div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown("""<div style="font-size:.65rem;color:#6b7a99;text-transform:uppercase;letter-spacing:2px;margin-bottom:.5rem;">⚡ Déficit / Superávit calórico</div>""", unsafe_allow_html=True)
    deficit_saved = int(gp.get("deficit", 0))
    deficit_val   = st.slider("kcal ajuste diario", min_value=-800, max_value=500, value=deficit_saved, step=50)
    meta_cal = limits["calorias"] + deficit_val
    if deficit_val < 0:
        d_color, d_label, d_tip = "#4facfe", f"Déficit de {abs(deficit_val)} kcal", "Ideal para pérdida de grasa"
    elif deficit_val > 0:
        d_color, d_label, d_tip = "#ff6b35", f"Superávit de {deficit_val} kcal", "Ideal para ganar masa muscular"
    else:
        d_color, d_label, d_tip = "#00e5a0", "Mantenimiento", "Sin ajuste calórico"

    st.markdown(f"""
    <div style="background:#1a2332;border:1px solid #1e2d45;border-radius:10px;padding:.8rem;margin-top:.3rem;text-align:center;">
        <div style="font-family:'JetBrains Mono',monospace;font-size:1.4rem;color:{d_color};font-weight:700;">{meta_cal} kcal</div>
        <div style="font-size:.72rem;color:{d_color};font-weight:600;margin-top:.2rem;">{d_label}</div>
        <div style="font-size:.65rem;color:#6b7a99;margin-top:.1rem;">{d_tip}</div>
        <div style="font-size:.6rem;color:#6b7a99;margin-top:.3rem;">Base: {limits['calorias']} kcal · Ajuste: {"+" if deficit_val>=0 else ""}{deficit_val}</div>
    </div>
    """, unsafe_allow_html=True)

    if deficit_val != deficit_saved:
        new_gp = dict(gp)
        new_gp["deficit"] = deficit_val
        sb_save_perfil(current_user, new_gp)
        st.session_state.perfil_cache["deficit"] = deficit_val

    limits["calorias"] = meta_cal

    st.markdown("<hr class='section-sep'>", unsafe_allow_html=True)
    bmi = weight / ((height / 100) ** 2)
    if bmi < 18.5:   bmi_label, bmi_color = "Bajo peso", "#4facfe"
    elif bmi < 25:   bmi_label, bmi_color = "Normal ✅",  "#00e5a0"
    elif bmi < 30:   bmi_label, bmi_color = "Sobrepeso",  "#ffd166"
    else:            bmi_label, bmi_color = "Obesidad",   "#ff4757"

    st.markdown(f"""<div style="background:#1a2332;border:1px solid #1e2d45;border-radius:12px;padding:1rem;text-align:center;"><div style="font-size:.65rem;color:#6b7a99;letter-spacing:2px;text-transform:uppercase;">IMC</div><div style="font-family:'Bebas Neue',sans-serif;font-size:2.5rem;color:{bmi_color};line-height:1.1;">{bmi:.1f}</div><div style="font-size:.8rem;color:{bmi_color};">{bmi_label}</div></div>""", unsafe_allow_html=True)

    st.markdown("<hr class='section-sep'>", unsafe_allow_html=True)
    if st.button("🚪 Cerrar Sesión", use_container_width=True):
        st.session_state.logged_in    = False
        st.session_state.username     = ""
        st.session_state.perfil_cache = {}
        st.session_state.editing_food_idx = None
        st.rerun()

# ─── NAVIGATION ────────────────────────────────────────────────────────────────
tab_nutricion, tab_ejercicio, tab_resumen, tab_database, tab_chat = st.tabs([
    "🥗 Nutrición", "🏋️ Rutina", "📊 Resumen", "🗄️ Database", "🤖 AI Coach"
])

# ═══════════════════════════════════════════════════════════════════════════════
# TAB 1 · NUTRICIÓN
# ═══════════════════════════════════════════════════════════════════════════════
with tab_nutricion:
    today_key = get_today_key()

    st.markdown(f"""<div class="card"><div class="card-title">🥗 REGISTRO DE ALIMENTOS</div><div style="font-size:.8rem;color:#6b7a99;">📅 Hoy: {date.today().strftime('%A %d de %B, %Y')}</div></div>""", unsafe_allow_html=True)

    with st.expander("➕ Agregar alimento", expanded=True):
        fk = st.session_state.food_form_counter
        c1, c2 = st.columns([2, 1])
        with c1:
            food_name = st.text_input("🍽️ Nombre del alimento", placeholder="Ej: Pechuga de pollo a la plancha", key=f"fn_{fk}")
            meal_type = st.selectbox("🕐 Momento del día", ["🌅 Desayuno","🌞 Almuerzo","🌙 Cena","🍎 Snack"], key=f"mt_{fk}")
        with c2:
            portion = st.text_input("📏 Porción", placeholder="200g / 1 taza", key=f"po_{fk}")

        st.markdown("**Información nutricional:**")
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            cal  = st.number_input("🔥 Calorías (kcal)", min_value=0.0, step=1.0, key=f"ca_{fk}")
            prot = st.number_input("💪 Proteínas (g)",   min_value=0.0, step=0.1, key=f"pr_{fk}")
        with col2:
            carbs = st.number_input("🌾 Carbos (g)",  min_value=0.0, step=0.1, key=f"cb_{fk}")
            fat   = st.number_input("🧈 Grasas (g)",  min_value=0.0, step=0.1, key=f"fa_{fk}")
        with col3:
            sugar = st.number_input("🍬 Azúcar (g)", min_value=0.0, step=0.1, key=f"su_{fk}")
            fiber = st.number_input("🌿 Fibra (g)",  min_value=0.0, step=0.1, key=f"fi_{fk}")
        with col4:
            sodium = st.number_input("🧂 Sodio (mg)",      min_value=0.0, step=1.0, key=f"so_{fk}")
            chol   = st.number_input("🫀 Colesterol (mg)", min_value=0.0, step=1.0, key=f"ch_{fk}")

        if st.button("✅ Agregar Alimento", use_container_width=True):
            if food_name.strip():
                food_record = {
                    "nombre": food_name, "porcion": portion, "comida": meal_type,
                    "calorias": cal, "proteinas": prot, "carbohidratos": carbs,
                    "grasas": fat, "azucar": sugar, "fibra": fiber,
                    "sodio": sodium, "colesterol": chol,
                    "hora": datetime.now().strftime("%H:%M"),
                }
                sb_add_alimento(current_user, today_key, food_record)
                st.session_state.food_form_counter += 1
                st.success(f"✅ '{food_name}' agregado")
                st.rerun()
            else:
                st.error("Ingresa el nombre del alimento")

    st.markdown("<hr class='section-sep'>", unsafe_allow_html=True)

    foods = sb_get_alimentos(current_user, today_key)

    if foods:
        totals = sum_nutrients(foods)

        st.markdown("""<div style="font-family:'Bebas Neue',sans-serif;font-size:1rem;letter-spacing:2px;color:#4facfe;margin-bottom:.8rem;">📊 TOTALES DEL DÍA</div>""", unsafe_allow_html=True)
        cols = st.columns(4)
        for i, (label, val, lim, unit) in enumerate([
            ("🔥 Calorías",  totals["calorias"],      limits["calorias"],      "kcal"),
            ("💪 Proteínas", totals["proteinas"],     limits["proteinas"],     "g"),
            ("🌾 Carbos",    totals["carbohidratos"], limits["carbohidratos"], "g"),
            ("🧈 Grasas",    totals["grasas"],        limits["grasas"],        "g"),
        ]):
            with cols[i]:
                pct   = min(val / lim * 100, 100) if lim > 0 else 0
                color = "#00e5a0" if pct<=80 else "#ffd166" if pct<=100 else "#ff4757"
                st.markdown(f"""<div class="metric-box"><div style="font-size:.65rem;color:#6b7a99;letter-spacing:1.5px;text-transform:uppercase;">{label}</div><div style="font-family:'Bebas Neue',sans-serif;font-size:1.8rem;color:{color};line-height:1.1;">{val:.0f}<span style="font-size:.9rem;">{unit}</span></div><div style="font-size:.65rem;color:#6b7a99;">/ {lim}{unit}</div><div style="margin-top:.5rem;height:4px;background:#1e2d45;border-radius:99px;"><div style="width:{pct:.0f}%;height:100%;border-radius:99px;background:{color};"></div></div></div>""", unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown("""<div style="font-family:'Bebas Neue',sans-serif;font-size:.95rem;letter-spacing:2px;color:#4facfe;margin-bottom:.8rem;">🔍 DETALLE NUTRICIONAL</div>""", unsafe_allow_html=True)

        col_a, col_b = st.columns(2)
        for i, (key, label, unit) in enumerate([
            ("azucar","🍬 Azúcar","g"), ("fibra","🌿 Fibra","g"),
            ("sodio","🧂 Sodio","mg"), ("colesterol","🫀 Colesterol","mg"),
        ]):
            val = totals[key]; lim = limits[key]
            pct = min(val/lim*100, 150) if lim>0 else 0
            status_txt, cls, msg = get_status(val, lim, key)
            bar_class = "prog-good" if cls=="good" else "prog-warn" if cls=="warn" else "prog-bad"
            target = col_a if i%2==0 else col_b
            with target:
                st.markdown(f"""<div class="prog-wrap"><div class="prog-label"><span>{label} <span class="pill pill-{cls}">{status_txt}</span></span><span>{val:.0f}/{lim}{unit}</span></div><div class="prog-bar"><div class="prog-fill {bar_class}" style="width:{min(pct,100):.0f}%"></div></div><div style="font-size:.68rem;color:#6b7a99;margin-top:.2rem;">{msg}</div></div>""", unsafe_allow_html=True)

        st.markdown("<hr class='section-sep'>", unsafe_allow_html=True)
        result = overall_diet_status(foods, limits)
        if result:
            cls, msg = result
            st.markdown(f'<div class="alert-{cls}"><strong>{msg}</strong></div>', unsafe_allow_html=True)

        st.markdown("<hr class='section-sep'>", unsafe_allow_html=True)
        st.markdown("""<div style="font-family:'Bebas Neue',sans-serif;font-size:.95rem;letter-spacing:2px;color:#00e5a0;margin-bottom:.8rem;">🍽️ ALIMENTOS REGISTRADOS</div>""", unsafe_allow_html=True)

        for i, food in enumerate(foods):
            is_editing = (st.session_state.editing_food_idx == i)
            col_f, col_btns = st.columns([11, 1])
            with col_f:
                prot_v = float(food.get("proteinas",0) or 0)
                carb_v = float(food.get("carbohidratos",0) or 0)
                gras_v = float(food.get("grasas",0) or 0)
                macros = f"P:{prot_v:.0f}g  C:{carb_v:.0f}g  G:{gras_v:.0f}g" if (prot_v+carb_v+gras_v)>0 else ""
                parts  = [p for p in [food.get("porcion","").strip(), macros, f"⏰ {food.get('hora','—')}"] if p]
                border = "#4facfe" if is_editing else "var(--border)"
                st.markdown(f"""<div class="food-row" style="border-color:{border};"><div><div class="food-name">{food['comida']} &nbsp; {food['nombre']}</div><div style="font-size:.72rem;color:#6b7a99;margin-top:.15rem;">{"  |  ".join(parts)}</div></div><div class="food-cal">{food['calorias']:.0f} kcal</div></div>""", unsafe_allow_html=True)
            with col_btns:
                edit_label = "✏️" if not is_editing else "✖️"
                if st.button(edit_label, key=f"edit_food_{i}", use_container_width=True):
                    st.session_state.editing_food_idx = None if is_editing else i
                    st.rerun()
                if st.button("🗑️", key=f"del_food_{i}", use_container_width=True):
                    sb_delete_alimento(food["id"])
                    if st.session_state.editing_food_idx == i:
                        st.session_state.editing_food_idx = None
                    st.rerun()

            if is_editing:
                st.markdown(f"""<div style="background:#0d1b2e;border:1px solid #4facfe;border-radius:12px;padding:1rem;margin:.2rem 0 .5rem 0;"><div style="font-family:'Bebas Neue',sans-serif;letter-spacing:2px;color:#4facfe;font-size:.85rem;margin-bottom:.7rem;">✏️ EDITANDO: {food['nombre']}</div></div>""", unsafe_allow_html=True)
                ec1, ec2 = st.columns([2,1])
                with ec1:
                    e_name = st.text_input("🍽️ Nombre", value=food.get("nombre",""), key=f"e_name_{i}")
                    meal_opts = ["🌅 Desayuno","🌞 Almuerzo","🌙 Cena","🍎 Snack"]
                    e_meal = st.selectbox("🕐 Momento", meal_opts,
                        index=meal_opts.index(food.get("comida","🌅 Desayuno")) if food.get("comida") in meal_opts else 0,
                        key=f"e_meal_{i}")
                with ec2:
                    e_portion = st.text_input("📏 Porción", value=food.get("porcion",""), key=f"e_portion_{i}")
                ea,eb,ec_,ed = st.columns(4)
                with ea:
                    e_cal  = st.number_input("🔥 Kcal",   value=float(food.get("calorias",0) or 0),  min_value=0.0, step=1.0,  key=f"e_cal_{i}")
                    e_prot = st.number_input("💪 Prot g",  value=float(food.get("proteinas",0) or 0), min_value=0.0, step=0.1,  key=f"e_prot_{i}")
                with eb:
                    e_carb = st.number_input("🌾 Carbs g", value=float(food.get("carbohidratos",0) or 0), min_value=0.0, step=0.1, key=f"e_carb_{i}")
                    e_fat  = st.number_input("🧈 Grasas g", value=float(food.get("grasas",0) or 0),  min_value=0.0, step=0.1,  key=f"e_fat_{i}")
                with ec_:
                    e_sug  = st.number_input("🍬 Azúcar g", value=float(food.get("azucar",0) or 0),  min_value=0.0, step=0.1,  key=f"e_sug_{i}")
                    e_fib  = st.number_input("🌿 Fibra g",  value=float(food.get("fibra",0) or 0),   min_value=0.0, step=0.1,  key=f"e_fib_{i}")
                with ed:
                    e_sod  = st.number_input("🧂 Sodio mg", value=float(food.get("sodio",0) or 0),   min_value=0.0, step=1.0,  key=f"e_sod_{i}")
                    e_cho  = st.number_input("🫀 Col mg",   value=float(food.get("colesterol",0) or 0), min_value=0.0, step=1.0, key=f"e_cho_{i}")
                if st.button("💾 Guardar cambios", key=f"save_edit_{i}", use_container_width=True):
                    updated = {
                        "nombre": e_name, "porcion": e_portion, "comida": e_meal,
                        "calorias": e_cal, "proteinas": e_prot, "carbohidratos": e_carb,
                        "grasas": e_fat, "azucar": e_sug, "fibra": e_fib,
                        "sodio": e_sod, "colesterol": e_cho,
                        "editado": datetime.now().strftime("%H:%M"),
                    }
                    sb_update_alimento(food["id"], updated)
                    st.session_state.editing_food_idx = None
                    st.rerun()
    else:
        st.markdown("""<div style="text-align:center;padding:3rem;color:#6b7a99;"><div style="font-size:3rem;margin-bottom:1rem;">🍽️</div><div>Sin alimentos registrados hoy</div></div>""", unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════════════
# TAB 2 · EJERCICIOS
# ═══════════════════════════════════════════════════════════════════════════════
with tab_ejercicio:
    st.markdown("""<div class="card card-orange"><div class="card-title">🏋️ MI RUTINA SEMANAL</div>
    <div style="font-size:.8rem;color:#6b7a99;">Tu rutina personalizada · Lunes a Sábado · Domingo = Descanso 💤</div>
    </div>""", unsafe_allow_html=True)

    rutina_usuario = get_rutina_usuario(current_user)
    dias_rutina    = ["Día 1","Día 2","Día 3","Día 4","Día 5","Día 6"]

    day_cols = st.columns(6)
    for i, dia_key in enumerate(dias_rutina):
        with day_cols[i]:
            dow_today = date.today().weekday()
            is_today  = (DIA_SEMANA_MAP.get(dow_today) == dia_key)
            label     = f"{'📍 ' if is_today else ''}{dia_key}"
            if st.button(label, key=f"daybtn_{i}", use_container_width=True):
                st.session_state.selected_day = dia_key

    selected      = st.session_state.selected_day
    dia_data_usr  = rutina_usuario.get(selected, {"titulo": "", "ejercicios": []})
    titulo_usr    = dia_data_usr.get("titulo", "")
    ejercicios_usr = list(dia_data_usr.get("ejercicios", []))

    st.markdown(f"""<div style="display:flex;align-items:center;gap:1rem;margin:1rem 0 .5rem;">
        <div style="font-family:'Bebas Neue',sans-serif;font-size:1.4rem;letter-spacing:3px;color:#ff6b35;">
            {selected.upper()}{(": " + titulo_usr.upper()) if titulo_usr else ""}
        </div>
    </div>""", unsafe_allow_html=True)

    with st.expander("📋 Ver / Editar rutina del día", expanded=True):
        if ejercicios_usr:
            for idx, ej in enumerate(ejercicios_usr):
                mc = {"Abdomen":"#ffd166","Espalda":"#00e5a0","Pecho":"#ff6b35",
                      "Pierna":"#4facfe","Bíceps":"#a78bfa","Triceps":"#ec4899",
                      "Cardio":"#ff4757","Core":"#ffd166","Espalda/Brazos":"#00e5a0",
                      "Hombros":"#4facfe","Glúteos":"#ec4899"}.get(ej.get("musculo",""),"#6b7a99")
                series_str = str(ej.get("series","-")) if ej.get("series",0) > 0 else "-"
                desc = ej.get("descanso", ej.get("descanso_txt", "-"))

                col_ex, col_actions = st.columns([10, 1])
                with col_ex:
                    st.markdown(f"""<div class="db-ex-row" style="padding:.5rem .3rem;">
                        <span style="color:{mc};font-weight:600;min-width:90px;font-size:.78rem;">{ej.get('musculo','—')}</span>
                        <span style="flex:1;color:#e8f0fe;font-size:.85rem;">{ej.get('nombre','—')}</span>
                        <span style="color:#6b7a99;font-family:'JetBrains Mono',monospace;font-size:.72rem;">
                            {series_str} series · {ej.get('reps','-')} · {desc}
                        </span>
                    </div>""", unsafe_allow_html=True)
                with col_actions:
                    subcols = st.columns(3)
                    with subcols[0]:
                        if st.button("↑", key=f"up_{selected}_{idx}", disabled=(idx==0), help="Subir"):
                            ejercicios_usr[idx], ejercicios_usr[idx-1] = ejercicios_usr[idx-1], ejercicios_usr[idx]
                            sb_save_rutina_dia(current_user, selected, titulo_usr, ejercicios_usr)
                            st.rerun()
                    with subcols[1]:
                        if st.button("↓", key=f"dn_{selected}_{idx}", disabled=(idx==len(ejercicios_usr)-1), help="Bajar"):
                            ejercicios_usr[idx], ejercicios_usr[idx+1] = ejercicios_usr[idx+1], ejercicios_usr[idx]
                            sb_save_rutina_dia(current_user, selected, titulo_usr, ejercicios_usr)
                            st.rerun()
                    with subcols[2]:
                        if st.button("✕", key=f"rdel_{selected}_{idx}", help="Eliminar"):
                            ejercicios_usr.pop(idx)
                            sb_save_rutina_dia(current_user, selected, titulo_usr, ejercicios_usr)
                            st.rerun()
        else:
            st.markdown("""<div style="text-align:center;padding:2rem;color:#6b7a99;">
                <div style="font-size:2rem;margin-bottom:.5rem;">🏗️</div>
                <div style="color:#a78bfa;font-weight:600;">Este día está vacío</div>
                <div style="font-size:.8rem;margin-top:.3rem;">Agrega ejercicios abajo</div>
            </div>""", unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)
        col_save, col_clear = st.columns(2)
        with col_save:
            if st.button(f"✅ Guardar rutina de hoy ({selected})", use_container_width=True, key="save_routine"):
                sb_clear_ejercicios(current_user, selected)
                for ej in ejercicios_usr:
                    sb_add_ejercicio(current_user, selected, {
                        "nombre":       ej.get("nombre",""),
                        "musculo":      ej.get("musculo",""),
                        "tipo":         "Fuerza",
                        "duracion":     0,
                        "series":       ej.get("series", 0),
                        "reps":         ej.get("reps",""),
                        "peso":         0.0,
                        "descanso_txt": ej.get("descanso", ej.get("descanso_txt","")),
                        "notas":        "",
                        "hora":         datetime.now().strftime("%H:%M"),
                        "fuente":       "rutina",
                    })
                st.success(f"✅ Rutina de {selected} guardada ({len(ejercicios_usr)} ejercicios)")
                st.rerun()
        with col_clear:
            if st.button("🗑️ Borrar ejercicios del día", use_container_width=True, key="clear_exercises"):
                sb_clear_ejercicios(current_user, selected)
                st.rerun()

    st.markdown("<hr class='section-sep'>", unsafe_allow_html=True)

    with st.expander("➕ Agregar ejercicio a mi rutina"):
        ec1, ec2 = st.columns([2,1])
        with ec1:
            ex_name  = st.text_input("🏋️ Nombre del ejercicio", placeholder="Ej: Press de banca...", key="ex_nombre")
            ex_group = st.selectbox("💪 Grupo muscular", [
                "Abdomen","Pecho","Espalda","Hombros","Bíceps","Triceps",
                "Pierna","Glúteos","Core","Cardio","Espalda/Brazos"
            ], key="ex_grupo")
        with ec2:
            ex_series  = st.number_input("📋 Series", min_value=0, max_value=20, value=3, key="ex_series")
            ex_reps_tx = st.text_input("🔁 Reps", placeholder="10, 10, 8", key="ex_reps")
            ex_rest    = st.text_input("😴 Descanso", placeholder="30 seg / 1 min", key="ex_descanso")

        if st.button("✅ Agregar a mi rutina", use_container_width=True, key="add_to_rutina"):
            if ex_name.strip():
                ejercicios_usr.append({
                    "musculo":  ex_group,
                    "nombre":   ex_name,
                    "series":   ex_series,
                    "reps":     ex_reps_tx,
                    "descanso": ex_rest,
                })
                sb_save_rutina_dia(current_user, selected, titulo_usr, ejercicios_usr)
                st.success(f"✅ '{ex_name}' agregado al {selected}")
                st.rerun()
            else:
                st.error("Ingresa el nombre del ejercicio")

    with st.expander("✏️ Editar título del día"):
        nuevo_titulo = st.text_input("Título", value=titulo_usr, key="titulo_dia",
                                     placeholder="Ej: Espalda, Pierna, Bíceps...")
        if st.button("💾 Guardar título", use_container_width=True, key="save_titulo"):
            sb_save_rutina_dia(current_user, selected, nuevo_titulo, ejercicios_usr)
            st.success("✅ Título actualizado")
            st.rerun()

    st.markdown("<hr class='section-sep'>", unsafe_allow_html=True)

    with st.expander("➕ Agregar ejercicio al registro de hoy (extra)"):
        ec1b, ec2b = st.columns([2,1])
        with ec1b:
            ex_name2  = st.text_input("🏋️ Nombre", placeholder="Ej: Press de banca...", key="ex2_nombre")
            ex_group2 = st.selectbox("💪 Grupo muscular", [
                "Pecho","Espalda","Hombros","Bíceps","Tríceps","Piernas",
                "Glúteos","Core/Abdomen","Cardio","Cuerpo completo"
            ], key="ex2_grupo")
        with ec2b:
            ex_type2 = st.selectbox("📌 Tipo", ["Fuerza","Cardio","Flexibilidad","HIIT","Funcional"], key="ex2_tipo")
            ex_sets2 = st.number_input("📋 Series", min_value=0, max_value=20, value=0, key="ex2_series")
            ex_reps2 = st.text_input("🔁 Reps", placeholder="10, 10, 8", key="ex2_reps")
            ex_rest2 = st.text_input("😴 Descanso", placeholder="30 seg", key="ex2_desc")
        ex_notes2 = st.text_area("📝 Notas", placeholder="RPE, sensaciones...", height=60, key="ex2_notas")

        if st.button("✅ Agregar al registro", use_container_width=True, key="add_custom_ex"):
            if ex_name2.strip():
                sb_add_ejercicio(current_user, selected, {
                    "nombre":       ex_name2,
                    "musculo":      ex_group2,
                    "tipo":         ex_type2,
                    "duracion":     0,
                    "series":       ex_sets2,
                    "reps":         ex_reps2,
                    "peso":         0.0,
                    "descanso_txt": ex_rest2,
                    "notas":        ex_notes2,
                    "hora":         datetime.now().strftime("%H:%M"),
                    "fuente":       "manual",
                })
                st.success(f"✅ '{ex_name2}' agregado al registro")
                st.rerun()
            else:
                st.error("Ingresa el nombre del ejercicio")

    exercises = sb_get_ejercicios(current_user, selected)

    if exercises:
        total_sets = sum(int(e.get("series",0) or 0) for e in exercises)
        grupos_uq  = len(set(get_musculo_from_exercise(e) for e in exercises))
        sc1,sc2,sc3 = st.columns(3)
        for col, val, label in zip([sc1,sc2,sc3],
                                   [len(exercises), total_sets, grupos_uq],
                                   ["Ejercicios","Series totales","Grupos musculares"]):
            with col:
                st.markdown(f"""<div class="metric-box"><div class="metric-val" style="color:#ff6b35;">{val}</div><div class="metric-label">{label}</div></div>""", unsafe_allow_html=True)

        st.markdown(f"""<br><div style="font-family:'Bebas Neue',sans-serif;font-size:.95rem;letter-spacing:2px;color:#ff6b35;margin-bottom:.8rem;">✅ EJERCICIOS REGISTRADOS HOY ({len(exercises)})</div>""", unsafe_allow_html=True)
        icon_map = {"Fuerza":"🏋️","Cardio":"🏃","Flexibilidad":"🧘","HIIT":"⚡","Funcional":"🤸"}

        for i, ex in enumerate(exercises):
            icon    = icon_map.get(ex.get("tipo","Fuerza"), "💪")
            musculo = get_musculo_from_exercise(ex)
            fuente_badge = '<span class="routine-badge" style="font-size:.6rem;">📋 Rutina</span>' if ex.get("fuente")=="rutina" else '<span class="pill pill-purple" style="font-size:.6rem;">✏️ Manual</span>'
            parts = []
            if ex.get("series"): parts.append(f"{ex['series']} series")
            if ex.get("reps"):   parts.append(str(ex["reps"]))
            if ex.get("peso"):   parts.append(f"{ex['peso']}kg")
            desc = ex.get("descanso_txt","") or ""
            if desc: parts.append(f"Desc: {desc}")
            detail = "  |  ".join(parts) if parts else "—"

            col_ex, col_del = st.columns([10,1])
            with col_ex:
                st.markdown(f"""<div class="ex-entry"><div class="ex-icon">{icon}</div><div class="ex-info"><div class="ex-name">{ex['nombre']} <span style="font-size:.7rem;color:#6b7a99;font-weight:400;">[{musculo}]</span> {fuente_badge}</div><div class="ex-detail">{detail}</div>{f'<div class="ex-detail" style="color:#4facfe;">📝 {ex["notas"]}</div>' if ex.get("notas") else ""}</div><div style="font-size:.7rem;color:#6b7a99;">⏰ {ex.get('hora','—')}</div></div>""", unsafe_allow_html=True)
            with col_del:
                if st.button("🗑️", key=f"del_ex_{selected}_{i}"):
                    sb_delete_ejercicio(ex["id"])
                    st.rerun()

        st.markdown("<hr class='section-sep'>", unsafe_allow_html=True)
        st.markdown("""<div style="font-family:'Bebas Neue',sans-serif;font-size:.9rem;letter-spacing:2px;color:#a78bfa;margin-bottom:.5rem;">🗄️ GUARDAR EN BASE DE DATOS</div>""", unsafe_allow_html=True)
        st.markdown("""<div class="alert-info">💡 Guarda el día completo (ejercicios + nutrición) en el historial.</div>""", unsafe_allow_html=True)

        coldb1, coldb2 = st.columns([2,1])
        with coldb1:
            suggested_week   = sb_get_next_semana_num(current_user)
            semana_num_input = st.number_input("Número de semana", min_value=1, max_value=52, value=suggested_week, step=1)
            dia_options      = ["DÍA 1","DÍA 2","DÍA 3","DÍA 4","DÍA 5","DÍA 6","DÍA 7"]
            suggested_dia_idx = dias_rutina.index(selected) if selected in dias_rutina else 0
            dia_label_input  = st.selectbox("Día de la semana", dia_options, index=suggested_dia_idx)
            notas_db = st.text_input("Notas del día (opcional)", placeholder="Ej: Buen entrenamiento...")
        with coldb2:
            st.markdown("<br><br>", unsafe_allow_html=True)
            today_foods_count = len(sb_get_alimentos(current_user, get_today_key()))
            st.markdown(f"""<div style="background:#1a2332;border:1px solid #1e2d45;border-radius:10px;padding:.8rem;font-size:.8rem;color:#6b7a99;"><div>💪 <strong style="color:#ff6b35;">{len(exercises)}</strong> ejercicios</div><div>🥗 <strong style="color:#00e5a0;">{today_foods_count}</strong> alimentos</div><div style="margin-top:.3rem;font-size:.7rem;">Se guardarán ambos</div></div>""", unsafe_allow_html=True)

        if st.button("💾 Guardar en Database", use_container_width=True, key="save_to_db"):
            today_foods_db = sb_get_alimentos(current_user, get_today_key())
            ex_clean   = [{k: v for k, v in e.items() if k not in ("id","username","created_at")} for e in exercises]
            food_clean = [{k: v for k, v in f.items() if k not in ("id","username","created_at")} for f in today_foods_db]
            sb_save_database(current_user, semana_num_input, dia_label_input, ex_clean, food_clean, notas_db)
            st.success(f"✅ Guardado en Semana {semana_num_input} · {dia_label_input}")

    else:
        st.markdown(f"""<div style="text-align:center;padding:2rem;color:#6b7a99;"><div style="font-size:3rem;margin-bottom:1rem;">🏋️</div><div>Sin ejercicios registrados para {selected}</div></div>""", unsafe_allow_html=True)

    st.markdown("""<div style="background:rgba(79,172,254,.08);border:1px solid rgba(79,172,254,.2);border-radius:12px;padding:1rem 1.2rem;margin-top:1rem;text-align:center;color:#4facfe;font-size:.85rem;">☀️ <strong>Domingo</strong> — Día de descanso y recuperación. 💤</div>""", unsafe_allow_html=True)

    # ── RESTAURAR ───────────────────────────────────────────────────────────────
    if current_user == "StephanoEl":
        st.markdown("<hr class='section-sep'>", unsafe_allow_html=True)
        with st.expander("⚙️ Admin · Restaurar rutina", expanded=False):
            st.markdown("""<div class="alert-warn">⚠️ Restaurar vuelve a la rutina original. No se puede deshacer.</div>""", unsafe_allow_html=True)
            col_rest1, col_rest2 = st.columns(2)
            with col_rest1:
                if st.button(f"🔄 Restaurar {selected}", use_container_width=True, key="r_restore_dia"):
                    default_dia = RUTINA_BASE.get(selected, {})
                    sb_save_rutina_dia(current_user, selected, default_dia.get("titulo",""), default_dia.get("ejercicios",[]))
                    st.success(f"✅ {selected} restaurado")
                    st.rerun()
            with col_rest2:
                if st.button("🔄 Restaurar toda la rutina", use_container_width=True, key="r_restore_all"):
                    for dk, dd in RUTINA_BASE.items():
                        sb_save_rutina_dia(current_user, dk, dd["titulo"], dd["ejercicios"])
                    st.success("✅ Rutina completa restaurada")
                    st.rerun()

with tab_resumen:
    st.markdown("""<div class="card card-blue"><div class="card-title">📊 RESUMEN DEL DÍA</div><div style="font-size:.8rem;color:#6b7a99;">Vista general de nutrición + entrenamiento de hoy</div></div>""", unsafe_allow_html=True)

    today_key_res   = get_today_key()
    foods_res       = sb_get_alimentos(current_user, today_key_res)
    dow_res         = date.today().weekday()
    today_dia_res   = DIA_SEMANA_MAP.get(dow_res)
    exercises_res   = sb_get_ejercicios(current_user, today_dia_res) if today_dia_res else []

    r1, r2 = st.columns(2)
    with r1:
        st.markdown("""<div style="font-family:'Bebas Neue',sans-serif;font-size:1rem;letter-spacing:2px;color:#00e5a0;margin-bottom:.8rem;">🥗 NUTRICIÓN HOY</div>""", unsafe_allow_html=True)
        if foods_res:
            totals  = sum_nutrients(foods_res)
            pct_cal = min(totals["calorias"]/limits["calorias"]*100, 100) if limits["calorias"]>0 else 0
            cal_col = "#00e5a0" if pct_cal<=80 else "#ffd166" if pct_cal<=100 else "#ff4757"
            st.markdown(f"""<div class="metric-box" style="margin-bottom:.8rem;"><div style="font-family:'Bebas Neue',sans-serif;font-size:2rem;color:{cal_col};">{totals['calorias']:.0f}</div><div class="metric-label">kcal consumidas de {limits['calorias']}</div></div>""", unsafe_allow_html=True)
            for key, label, unit in [
                ("proteinas","💪 Proteínas","g"),("carbohidratos","🌾 Carbos","g"),
                ("grasas","🧈 Grasas","g"),("azucar","🍬 Azúcar","g"),
                ("fibra","🌿 Fibra","g"),("sodio","🧂 Sodio","mg"),
            ]:
                val = totals[key]; lim = limits[key]
                pct = min(val/lim*100, 100) if lim>0 else 0
                st_txt,cls,_ = get_status(val,lim,key)
                bc = "prog-good" if cls=="good" else "prog-warn" if cls=="warn" else "prog-bad"
                st.markdown(f"""<div class="prog-wrap"><div class="prog-label"><span>{label} <span class="pill pill-{cls}">{st_txt}</span></span><span>{val:.0f}/{lim}{unit}</span></div><div class="prog-bar"><div class="prog-fill {bc}" style="width:{pct:.0f}%"></div></div></div>""", unsafe_allow_html=True)
            result = overall_diet_status(foods_res, limits)
            if result:
                cls,msg = result
                st.markdown(f'<div class="alert-{cls}" style="margin-top:.8rem;">{msg}</div>', unsafe_allow_html=True)
        else:
            st.markdown('<div style="color:#6b7a99;padding:2rem;text-align:center;">Sin datos de nutrición hoy</div>', unsafe_allow_html=True)

    with r2:
        today_dia_label = RUTINA_BASE.get(today_dia_res,{}).get("titulo","Domingo") if today_dia_res else "Domingo"
        st.markdown(f"""<div style="font-family:'Bebas Neue',sans-serif;font-size:1rem;letter-spacing:2px;color:#ff6b35;margin-bottom:.8rem;">🏋️ ENTRENAMIENTO HOY {"("+today_dia_label+")" if today_dia_res else "(DOMINGO)"}</div>""", unsafe_allow_html=True)
        if not today_dia_res:
            st.markdown("""<div style="text-align:center;padding:2rem;color:#4facfe;"><div style="font-size:3rem;">😴</div><div>Hoy es Domingo · Día de descanso</div></div>""", unsafe_allow_html=True)
        elif exercises_res:
            total_sets = sum(int(e.get("series",0) or 0) for e in exercises_res)
            grupos     = list(set(get_musculo_from_exercise(e) for e in exercises_res))
            st.markdown(f"""<div class="metric-box" style="margin-bottom:.8rem;"><div class="metric-val" style="color:#ff6b35;">{len(exercises_res)}</div><div class="metric-label">ejercicios · {total_sets} series</div></div>""", unsafe_allow_html=True)
            st.markdown(f"""<div style="font-size:.75rem;color:#6b7a99;margin-bottom:.5rem;">Grupos: {', '.join(grupos)}</div>""", unsafe_allow_html=True)
            for ex in exercises_res:
                parts = []
                if ex.get("series"): parts.append(f"{ex['series']} series")
                if ex.get("reps"):   parts.append(str(ex["reps"]))
                st.markdown(f"""<div style="background:#1a2332;border:1px solid #1e2d45;border-radius:8px;padding:.6rem .8rem;margin:.3rem 0;font-size:.82rem;"><strong>{ex['nombre']}</strong><span style="color:#6b7a99;margin-left:.5rem;">{' | '.join(parts)}</span></div>""", unsafe_allow_html=True)
        else:
            rutina_hoy = RUTINA_BASE.get(today_dia_res)
            if rutina_hoy:
                st.markdown(f"""<div class="alert-warn">⚡ Sin ejercicios registrados. Tu rutina de hoy: <strong>{rutina_hoy['titulo']}</strong> ({len(rutina_hoy['ejercicios'])} ejercicios).</div>""", unsafe_allow_html=True)

    st.markdown("<hr class='section-sep'>", unsafe_allow_html=True)
    st.markdown("""<div style="font-family:'Bebas Neue',sans-serif;font-size:1rem;letter-spacing:2px;color:#4facfe;margin-bottom:.8rem;">📅 RESUMEN SEMANAL</div>""", unsafe_allow_html=True)
    week_cols = st.columns(7)
    dia_names = ["Lun","Mar","Mié","Jue","Vie","Sáb","Dom"]
    for i in range(7):
        dia_key_w = DIA_SEMANA_MAP.get(i)
        exs_w     = sb_get_ejercicios(current_user, dia_key_w) if dia_key_w else []
        is_tod    = (date.today().weekday()==i)
        color     = "#ff6b35" if is_tod else "#4facfe" if exs_w else "#6b7a99"
        icon      = "📍" if is_tod else "✅" if exs_w else ("😴" if i==6 else "○")
        with week_cols[i]:
            st.markdown(f"""<div style="text-align:center;background:#111827;border:1px solid {'#ff6b35' if is_tod else '#1e2d45'};border-radius:10px;padding:.8rem .3rem;"><div style="font-size:1.2rem;">{icon}</div><div style="font-family:'Bebas Neue',sans-serif;letter-spacing:1px;color:{color};font-size:.9rem;">{dia_names[i]}</div><div style="font-size:.65rem;color:#6b7a99;margin-top:.2rem;">{len(exs_w)} ej.</div></div>""", unsafe_allow_html=True)
# ═══════════════════════════════════════════════════════════════════════════════
# TAB 4 · DATABASE
# ═══════════════════════════════════════════════════════════════════════════════
with tab_database:
    st.markdown("""<div class="card card-purple"><div class="card-title">🗄️ BASE DE DATOS HISTÓRICA</div><div style="font-size:.8rem;color:#6b7a99;">Historial completo por semana y día · Ejercicios + Nutrición + Notas</div></div>""", unsafe_allow_html=True)

    db_rows = sb_get_database(current_user)

    if not db_rows:
        st.markdown("""
        <div style="text-align:center;padding:4rem;color:#6b7a99;">
            <div style="font-size:4rem;margin-bottom:1rem;">🗄️</div>
            <div style="font-size:1.2rem;font-weight:600;color:#a78bfa;">Base de datos vacía</div>
            <div style="font-size:.9rem;margin-top:.8rem;">Ve a <strong style="color:#ff6b35;">Rutina</strong> y usa <strong>💾 Guardar en Database</strong>.</div>
        </div>""", unsafe_allow_html=True)
    else:
        # Agrupar por semana
        semanas = {}
        for row in db_rows:
            sk = f"Semana {row['semana_num']}"
            if sk not in semanas:
                semanas[sk] = []
            semanas[sk].append(row)

        total_sem          = len(semanas)
        total_dias         = len(db_rows)
        total_ex_global    = sum(len(r.get("ejercicios") or []) for r in db_rows)
        total_meals_global = sum(len(r.get("alimentos") or []) for r in db_rows)

        # Filtros
        fcol1, fcol2, fcol3 = st.columns(3)
        with fcol1:
            all_week_keys  = sorted(semanas.keys(), key=lambda x: int(x.split()[-1]))
            filter_week    = st.selectbox("🗓️ Filtrar por semana", ["Todas"] + all_week_keys, key="db_filter_week")
        with fcol2:
            filter_dia     = st.selectbox("📅 Filtrar por día", ["Todos","DÍA 1","DÍA 2","DÍA 3","DÍA 4","DÍA 5","DÍA 6","DÍA 7"], key="db_filter_dia")
        with fcol3:
            filter_tipo    = st.selectbox("🔍 Mostrar", ["Todo","Solo ejercicios","Solo nutrición"], key="db_filter_tipo")

        st.markdown("<hr class='section-sep'>", unsafe_allow_html=True)

        gs1, gs2, gs3, gs4 = st.columns(4)
        for col, val, label, color in zip(
            [gs1, gs2, gs3, gs4],
            [total_sem, total_dias, total_ex_global, total_meals_global],
            ["Semanas","Días registrados","Ejercicios totales","Comidas registradas"],
            ["#a78bfa","#4facfe","#ff6b35","#00e5a0"],
        ):
            with col:
                st.markdown(f"""<div class="metric-box"><div class="metric-val" style="color:{color};">{val}</div><div class="metric-label">{label}</div></div>""", unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)

        rows_filtered = [
            r for r in db_rows
            if (filter_week == "Todas" or f"Semana {r['semana_num']}" == filter_week)
            and (filter_dia == "Todos" or r["dia_label"].upper().replace(" ","") == filter_dia.upper().replace(" ",""))
        ]

        semanas_filtered = {}
        for row in rows_filtered:
            sk = f"Semana {row['semana_num']}"
            if sk not in semanas_filtered:
                semanas_filtered[sk] = []
            semanas_filtered[sk].append(row)

        for sem_key in sorted(semanas_filtered.keys(), key=lambda x: int(x.split()[-1]), reverse=True):
            st.markdown(f"""<div class="db-week-header">📅 {sem_key.upper()}</div>""", unsafe_allow_html=True)

            for row in sorted(semanas_filtered[sem_key], key=lambda x: x["dia_label"]):
                ejercicios = row.get("ejercicios") or []
                alimentos  = row.get("alimentos") or []
                notas      = row.get("notas") or ""
                fecha_dia  = row.get("fecha") or ""
                hora_g     = row.get("guardado_en") or ""

                try:
                    fecha_fmt = date.fromisoformat(fecha_dia).strftime("%d/%m/%Y (%A)")
                except Exception:
                    fecha_fmt = fecha_dia

                nut_totals = sum_nutrients(alimentos) if alimentos else {}
                cal_dia    = nut_totals.get("calorias", 0)
                ex_count   = len([e for e in ejercicios if e.get("nombre","") != "Descanso"])

                with st.expander(f"{'📍 ' if fecha_dia == get_today_key() else ''}{row['dia_label']} · {fecha_fmt} · {ex_count} ejercicios · {cal_dia:.0f} kcal"):
                    if notas:
                        st.markdown(f"""<div class="alert-info">📝 {notas}</div>""", unsafe_allow_html=True)

                    grupos_dia = list(set(get_musculo_from_exercise(e) for e in ejercicios if e.get("nombre","") != "Descanso"))
                    if grupos_dia:
                        badges = " ".join(f'<span class="pill pill-info">{g}</span>' for g in grupos_dia[:5])
                        st.markdown(f"<div style='margin-bottom:.8rem;'>{badges}</div>", unsafe_allow_html=True)

                    db_col1, db_col2 = st.columns(2)
                    with db_col1:
                        if filter_tipo != "Solo nutrición":
                            st.markdown("""<div style="font-family:'Bebas Neue',sans-serif;font-size:.9rem;letter-spacing:2px;color:#ff6b35;margin-bottom:.5rem;">🏋️ EJERCICIOS</div>""", unsafe_allow_html=True)
                            if ejercicios:
                                for ex in ejercicios:
                                    if ex.get("nombre") == "Descanso": continue
                                    musculo = get_musculo_from_exercise(ex)
                                    mc = {"Abdomen":"#ffd166","Espalda":"#00e5a0","Pecho":"#ff6b35","Pierna":"#4facfe","Bíceps":"#a78bfa","Triceps":"#ec4899"}.get(musculo, "#6b7a99")
                                    fuente = "📊" if ex.get("fuente") == "excel" else "✏️"
                                    st.markdown(f"""<div class="db-ex-row"><span style="color:{mc};font-weight:600;min-width:85px;font-size:.75rem;">{musculo}</span><span class="db-nombre">{fuente} {ex.get('nombre','')}</span><span class="db-detail">{ex.get('series','-')}s · {ex.get('reps','-')}</span></div>""", unsafe_allow_html=True)
                            else:
                                st.markdown('<div style="color:#6b7a99;font-size:.8rem;">Sin ejercicios</div>', unsafe_allow_html=True)

                    with db_col2:
                        if filter_tipo != "Solo ejercicios":
                            st.markdown("""<div style="font-family:'Bebas Neue',sans-serif;font-size:.9rem;letter-spacing:2px;color:#00e5a0;margin-bottom:.5rem;">🥗 NUTRICIÓN</div>""", unsafe_allow_html=True)
                            if alimentos:
                                prot_dia = nut_totals.get("proteinas", 0)
                                st.markdown(f"""<div style="background:#1a2332;border:1px solid #1e2d45;border-radius:8px;padding:.7rem;margin-bottom:.5rem;"><span style="font-family:'JetBrains Mono',monospace;color:#00e5a0;font-size:.85rem;">🔥 {cal_dia:.0f} kcal</span><span style="color:#6b7a99;font-size:.8rem;"> · P:{prot_dia:.0f}g · C:{nut_totals.get('carbohidratos',0):.0f}g · G:{nut_totals.get('grasas',0):.0f}g</span></div>""", unsafe_allow_html=True)
                                for food in alimentos:
                                    st.markdown(f"""<div class="db-ex-row"><span style="color:#6b7a99;min-width:70px;font-size:.72rem;">{str(food.get('comida','—'))[:10]}</span><span class="db-nombre" style="font-size:.8rem;">{food.get('nombre','—')}</span><span class="db-detail">{food.get('calorias',0):.0f} kcal</span></div>""", unsafe_allow_html=True)
                            else:
                                st.markdown('<div style="color:#6b7a99;font-size:.8rem;">Sin alimentos</div>', unsafe_allow_html=True)

                    st.markdown(f"""<div style="text-align:right;font-size:.65rem;color:#6b7a99;margin-top:.5rem;padding-top:.5rem;border-top:1px solid #1e2d45;">💾 Guardado a las {hora_g}</div>""", unsafe_allow_html=True)

                    if st.button(f"🗑️ Eliminar {row['dia_label']} de {sem_key}", key=f"del_db_{row['id']}"):
                        sb_delete_database_row(row["id"])
                        st.rerun()

        # ─── EXPORTAR EXCEL ───────────────────────────────────────────────────
        st.markdown("<hr class='section-sep'>", unsafe_allow_html=True)
        st.markdown("""<div style="font-family:'Bebas Neue',sans-serif;font-size:.9rem;letter-spacing:2px;color:#a78bfa;margin-bottom:.8rem;">📤 EXPORTAR DATOS</div>""", unsafe_allow_html=True)

        ex1, ex2, ex3 = st.columns(3)
        with ex1:
            exp_semana = st.selectbox("📅 Semana a exportar", ["Todas"] + sorted(semanas.keys(), key=lambda x: int(x.split()[-1])), key="exp_semana")
        with ex2:
            exp_dia = st.selectbox("📋 Día a exportar", ["Todos","DÍA 1","DÍA 2","DÍA 3","DÍA 4","DÍA 5","DÍA 6","DÍA 7"], key="exp_dia")
        with ex3:
            st.markdown("<br>", unsafe_allow_html=True)
            generar = st.button("⚙️ Generar Excel", use_container_width=True, key="generar_excel")

        if generar:
            filas_ej_exp = []
            filas_al_exp = []
            for row in db_rows:
                sem_key_check = f"Semana {row['semana_num']}"
                if exp_semana != "Todas" and sem_key_check != exp_semana:
                    continue
                if exp_dia != "Todos" and row["dia_label"].upper().replace(" ","") != exp_dia.upper().replace(" ",""):
                    continue
                fecha = row.get("fecha","")
                for ex in (row.get("ejercicios") or []):
                    if ex.get("nombre") != "Descanso":
                        filas_ej_exp.append({
                            "Semana": sem_key_check, "Día": row["dia_label"], "Fecha": fecha,
                            "Músculo":   get_musculo_from_exercise(ex),
                            "Ejercicio": ex.get("nombre",""),
                            "Series":    ex.get("series",""),
                            "Reps":      ex.get("reps",""),
                            "Peso (kg)": ex.get("peso",""),
                            "Descanso":  ex.get("descanso_txt",""),
                            "Notas":     ex.get("notas",""),
                        })
                for food in (row.get("alimentos") or []):
                    filas_al_exp.append({
                        "Semana": sem_key_check, "Día": row["dia_label"], "Fecha": fecha,
                        "Comida":          food.get("comida",""),
                        "Alimento":        food.get("nombre",""),
                        "Porción":         food.get("porcion",""),
                        "Calorías (kcal)": food.get("calorias",0),
                        "Proteínas (g)":   food.get("proteinas",0),
                        "Carbos (g)":      food.get("carbohidratos",0),
                        "Grasas (g)":      food.get("grasas",0),
                        "Azúcar (g)":      food.get("azucar",0),
                        "Fibra (g)":       food.get("fibra",0),
                        "Sodio (mg)":      food.get("sodio",0),
                        "Colesterol (mg)": food.get("colesterol",0),
                    })

            if not filas_ej_exp and not filas_al_exp:
                st.markdown('<div class="alert-warn">⚠️ No hay datos para los filtros seleccionados.</div>', unsafe_allow_html=True)
            else:
                import io
                from openpyxl import Workbook
                from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
                from openpyxl.utils import get_column_letter
                from openpyxl.formatting.rule import ColorScaleRule

                wb  = Workbook()
                C_WHITE = "FFFFFF"; C_LIGHT = "F8F9FC"; C_LIGHT2 = "F0F2F8"
                C_BORDER = "E2E6F0"; C_DARK = "2D3452"; C_MUTED = "8B93B0"
                C_ORANGE = "E8805A"; C_GREEN = "4CAF8A"; C_BLUE = "5B8DEF"
                C_PURPLE = "8B72BE"
                th_s = Side(style="thin", color=C_BORDER)
                tk_s = Side(style="medium", color="C8CEDF")
                thin_b  = Border(left=th_s, right=th_s, top=th_s, bottom=th_s)
                thick_b = Border(left=tk_s, right=tk_s, top=tk_s, bottom=tk_s)

                def sh(cell, bg, txt="FFFFFF", bold=True, size=11):
                    cell.fill = PatternFill("solid", fgColor=bg)
                    cell.font = Font(bold=bold, color=txt, size=size, name="Calibri")
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.border = thick_b

                def sc(cell, row_num, fg=None, bold=False, left=False, bg_override=None):
                    bg = bg_override or (C_LIGHT2 if row_num % 2 == 0 else C_WHITE)
                    cell.fill = PatternFill("solid", fgColor=bg)
                    cell.font = Font(color=fg or C_DARK, size=10, bold=bold, name="Calibri")
                    cell.alignment = Alignment(horizontal="left" if left else "center", vertical="center")
                    cell.border = thin_b

                def aw(ws):
                    for col in ws.columns:
                        mx = max((len(str(c.value or "")) for c in col), default=10)
                        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(mx+3,10),50)

                filtro_label = exp_semana + (f" · {exp_dia}" if exp_dia != "Todos" else "")

                # Portada
                wc = wb.active; wc.title = "Portada"
                wc.sheet_view.showGridLines = False
                for ci, w in enumerate([2,18,18,18,18,18,18,18,2], 1):
                    wc.column_dimensions[get_column_letter(ci)].width = w
                for r in range(1,7):
                    for c in range(1,10):
                        wc.cell(row=r,column=c).fill = PatternFill("solid", fgColor=C_ORANGE if r<=4 else "EF9E7A")
                    wc.row_dimensions[r].height = 18
                wc.merge_cells("B2:H4"); t = wc["B2"]
                t.value = "🔥  FIREMUSCLE"; t.font = Font(bold=True,size=32,color=C_WHITE,name="Calibri")
                t.alignment = Alignment(horizontal="center",vertical="center")
                t.fill = PatternFill("solid",fgColor=C_ORANGE)
                wc.merge_cells("B5:H6"); st2 = wc["B5"]
                st2.value = "REPORTE DE PROGRESO PERSONAL"
                st2.font = Font(size=13,color="FFE0D0",name="Calibri")
                st2.alignment = Alignment(horizontal="center",vertical="center")
                st2.fill = PatternFill("solid",fgColor="FF8050")
                wc.merge_cells("B8:H8"); inf = wc["B8"]
                inf.value = f"Usuario: {current_user}   ·   {date.today().strftime('%d de %B del %Y')}   ·   Filtro: {filtro_label}"
                inf.font = Font(size=10,color=C_MUTED,name="Calibri")
                inf.alignment = Alignment(horizontal="center",vertical="center")
                inf.fill = PatternFill("solid",fgColor=C_WHITE)
                wc.row_dimensions[8].height = 20

                card_data = [("Semanas",str(total_sem),C_PURPLE,"F3EFF9"),("Días",str(total_dias),C_BLUE,"EEF3FD"),("Ejercicios",str(len(filas_ej_exp)),C_ORANGE,"FDF0EB"),("Comidas",str(len(filas_al_exp)),C_GREEN,"EDF7F3")]
                for idx, ((lbl,val,clr,bg2),(col_s,col_e)) in enumerate(zip(card_data,zip(["B","D","F","H"],["C","E","G","I"]))):
                    wc.merge_cells(f"{col_s}11:{col_e}13"); vc = wc[f"{col_s}11"]
                    vc.value=val; vc.font=Font(bold=True,size=30,color=clr,name="Calibri")
                    vc.alignment=Alignment(horizontal="center",vertical="center")
                    vc.fill=PatternFill("solid",fgColor=bg2)
                    wc.merge_cells(f"{col_s}14:{col_e}14"); lc = wc[f"{col_s}14"]
                    lc.value=lbl; lc.font=Font(bold=True,size=10,color=clr,name="Calibri")
                    lc.alignment=Alignment(horizontal="center",vertical="center")
                    lc.fill=PatternFill("solid",fgColor=bg2)

                # Hoja ejercicios
                we = wb.create_sheet("Ejercicios"); we.sheet_view.showGridLines = False
                if filas_ej_exp:
                    df_e = pd.DataFrame(filas_ej_exp); ncols = len(df_e.columns)
                    we.merge_cells(f"A1:{get_column_letter(ncols)}1"); t1 = we["A1"]
                    t1.value="💪  REGISTRO DE EJERCICIOS"; t1.font=Font(bold=True,size=14,color=C_WHITE,name="Calibri")
                    t1.alignment=Alignment(horizontal="center",vertical="center")
                    t1.fill=PatternFill("solid",fgColor=C_ORANGE); we.row_dimensions[1].height=28
                    hc={"Semana":"A08CC8","Día":"A08CC8","Fecha":"A08CC8","Músculo":"E8956A","Ejercicio":"E8956A","Series":"7AAAE8","Reps":"7AAAE8","Peso (kg)":"6BBD9E","Descanso":"D4A843","Notas":"9AA3BE"}
                    for ci,h in enumerate(df_e.columns,1):
                        sh(we.cell(row=2,column=ci,value=h), hc.get(h,C_DARK))
                    we.row_dimensions[2].height=22
                    mc_map={"Abdomen":"C49A2A","Espalda":"3D9E7A","Pecho":"C86845","Pierna":"4A78D4","Bíceps":"7A5CAE","Triceps":"B85A8A"}
                    pill_bg={"Abdomen":"FBF6E9","Espalda":"EDF7F3","Pecho":"FBF0EB","Pierna":"EEF3FD","Bíceps":"F3EFF9","Triceps":"FAF0F6"}
                    for ri,row in enumerate(df_e.itertuples(index=False),3):
                        musc=str(row[3]) if len(row)>3 else ""
                        acc=mc_map.get(musc,C_DARK); pbg=pill_bg.get(musc,C_LIGHT)
                        for ci,v in enumerate(row,1):
                            cell=we.cell(row=ri,column=ci,value=v)
                            if ci in (4,5): sc(cell,ri-2,fg=acc,bold=(ci==5),left=(ci==5),bg_override=pbg)
                            else: sc(cell,ri-2)
                        we.row_dimensions[ri].height=18
                    aw(we)

                # Hoja nutrición
                wn = wb.create_sheet("Nutricion"); wn.sheet_view.showGridLines = False
                if filas_al_exp:
                    df_a = pd.DataFrame(filas_al_exp); ncols_a = len(df_a.columns)
                    wn.merge_cells(f"A1:{get_column_letter(ncols_a)}1"); t2 = wn["A1"]
                    t2.value="🥗  REGISTRO NUTRICIONAL"; t2.font=Font(bold=True,size=14,color=C_WHITE,name="Calibri")
                    t2.alignment=Alignment(horizontal="center",vertical="center")
                    t2.fill=PatternFill("solid",fgColor=C_GREEN); wn.row_dimensions[1].height=28
                    ha={"Semana":"A08CC8","Día":"A08CC8","Fecha":"A08CC8","Comida":"7AAAE8","Alimento":"6BBD9E","Porción":"9AA3BE","Calorías (kcal)":"E8956A","Proteínas (g)":"6BBD9E","Carbos (g)":"7AAAE8","Grasas (g)":"D4A843","Azúcar (g)":"D472A0","Fibra (g)":"4DADA0","Sodio (mg)":"5BAFD4","Colesterol (mg)":"D46B6B"}
                    for ci,h in enumerate(df_a.columns,1):
                        sh(wn.cell(row=2,column=ci,value=h), ha.get(h,C_DARK))
                    wn.row_dimensions[2].height=22
                    cc={"Desayuno":"C49A2A","Almuerzo":"C86845","Cena":"7A5CAE","Snack":"3D9E7A"}
                    cb={"Desayuno":"FBF6E9","Almuerzo":"FBF0EB","Cena":"F3EFF9","Snack":"EDF7F3"}
                    for ri,row in enumerate(df_a.itertuples(index=False),3):
                        cv=str(row[3]) if len(row)>3 else ""
                        acc=next((v for k,v in cc.items() if k in cv),C_DARK)
                        pbg=next((v for k,v in cb.items() if k in cv),C_LIGHT)
                        for ci,v in enumerate(row,1):
                            cell=wn.cell(row=ri,column=ci,value=v)
                            if ci in (4,5): sc(cell,ri-2,fg=acc,bold=(ci==5),left=(ci==5),bg_override=pbg)
                            else: sc(cell,ri-2)
                        wn.row_dimensions[ri].height=18
                    tr=len(df_a)+3; wn.row_dimensions[tr].height=22
                    for ci in range(1,ncols_a+1):
                        cell=wn.cell(row=tr,column=ci)
                        if ci==5:
                            cell.value="TOTALES"; cell.font=Font(bold=True,size=11,color=C_WHITE,name="Calibri")
                            cell.fill=PatternFill("solid",fgColor=C_GREEN)
                            cell.alignment=Alignment(horizontal="center",vertical="center"); cell.border=thick_b
                        elif ci>=7:
                            cl=get_column_letter(ci); cell.value=f"=SUM({cl}3:{cl}{tr-1})"
                            cell.number_format="0.0"; cell.font=Font(bold=True,color=C_GREEN,size=11,name="Calibri")
                            cell.fill=PatternFill("solid",fgColor="E6FAF4")
                            cell.alignment=Alignment(horizontal="center",vertical="center"); cell.border=thick_b
                        else:
                            cell.fill=PatternFill("solid",fgColor=C_LIGHT2); cell.border=thick_b
                    aw(wn)

                buf = io.BytesIO(); wb.save(buf); buf.seek(0)
                fname = f"firemuscle_{current_user}_{date.today().isoformat()}.xlsx"
                st.session_state["excel_buf"]   = buf
                st.session_state["excel_fname"] = fname
                st.success(f"✅ Excel listo — {len(filas_ej_exp)} ejercicios · {len(filas_al_exp)} alimentos")

        if "excel_buf" in st.session_state:
            st.download_button(
                label="📥 Descargar Excel",
                data=st.session_state["excel_buf"],
                file_name=st.session_state["excel_fname"],
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

# ═══════════════════════════════════════════════════════════════════════════════
# TAB 5 · AI COACH
# ═══════════════════════════════════════════════════════════════════════════════
with tab_chat:
    st.markdown("""<div class="card card-purple"><div class="card-title">🤖 AI COACH · FIREMUSCLE</div>
    <div style="font-size:.8rem;color:#6b7a99;">Asistente personal con acceso a tu perfil, rutina y nutrición</div>
    </div>""", unsafe_allow_html=True)

    today_foods_chat    = sb_get_alimentos(current_user, get_today_key())
    dow_chat            = date.today().weekday()
    today_dia_key_chat  = DIA_SEMANA_MAP.get(dow_chat)
    exercises_today_chat = sb_get_ejercicios(current_user, today_dia_key_chat) if today_dia_key_chat else []

    user_context = f"""
Eres el AI Coach personal de FireMuscle, un asistente experto en fitness y nutrición.
Tienes acceso al perfil completo del usuario y debes personalizar cada respuesta con sus datos.

PERFIL DEL USUARIO:
- Nombre: {current_user}
- Edad: {gp.get('age', '?')} años
- Peso: {gp.get('weight', '?')} kg
- Altura: {gp.get('height', '?')} cm
- Sexo: {gp.get('sex', '?')}
- Nivel de actividad: {gp.get('activity', '?')}
- Objetivos calóricos: {limits.get('calorias', '?')} kcal/día
- Proteínas meta: {limits.get('proteinas', '?')}g/día

RUTINA SEMANAL (6 días):
{chr(10).join([f"- {k}: {v['titulo']} ({len(v['ejercicios'])} ejercicios)" for k, v in RUTINA_BASE.items()])}

REGISTRO DE HOY:
- Día de entrenamiento: {RUTINA_BASE.get(today_dia_key_chat, {}).get('titulo', 'Domingo/Descanso')}
- Ejercicios registrados: {len(exercises_today_chat)}
- Alimentos registrados: {len(today_foods_chat)}
- Calorías consumidas: {sum_nutrients(today_foods_chat).get('calorias', 0):.0f} kcal de {limits.get('calorias', '?')} meta

Responde siempre en español. Sé directo, motivador y usa los datos del usuario para personalizar cada respuesta.
"""

    if "chat_messages" not in st.session_state:
        st.session_state.chat_messages = []

    st.markdown("""<div style="font-size:.75rem;color:#6b7a99;margin-bottom:.5rem;letter-spacing:1px;">💡 PREGUNTAS RÁPIDAS</div>""", unsafe_allow_html=True)
    quick_cols = st.columns(4)
    quick_prompts = [
        ("💪 ¿Cómo mejorar mi rutina?", "Analiza mi rutina actual y dime qué ejercicios podría mejorar o agregar"),
        ("🥗 ¿Cómo está mi nutrición?",  "Revisa mi alimentación de hoy y dime si estoy cumpliendo mis metas"),
        ("🔥 Rutina de emergencia",       "Necesito una rutina rápida de 20 minutos para hacer en casa hoy"),
        ("📈 ¿Cómo progresar?",           "Dame consejos específicos para seguir progresando según mi nivel"),
    ]
    for i, (label, prompt) in enumerate(quick_prompts):
        with quick_cols[i]:
            if st.button(label, key=f"quick_{i}", use_container_width=True):
                st.session_state.chat_messages.append({"role": "user", "content": prompt})

    st.markdown("<hr class='section-sep'>", unsafe_allow_html=True)

    for msg in st.session_state.chat_messages:
        role_label   = f"👤 {current_user}" if msg["role"] == "user" else "🔥 AI Coach"
        bubble_style = "background:#ff6b35;color:#fff;" if msg["role"] == "user" else "background:#1a2332;border:1px solid #1e2d45;color:#e8f0fe;"
        align        = "flex-end" if msg["role"] == "user" else "flex-start"
        st.markdown(f"""
        <div style="display:flex;justify-content:{align};margin:.4rem 0;">
            <div style="{bubble_style}border-radius:14px;padding:.7rem 1rem;max-width:80%;font-size:.88rem;line-height:1.5;">
                <div style="font-size:.65rem;opacity:.7;margin-bottom:.3rem;">{role_label}</div>
                {msg['content']}
            </div>
        </div>
        """, unsafe_allow_html=True)

    if st.session_state.chat_messages and st.session_state.chat_messages[-1]["role"] == "user":
        with st.spinner("🔥 AI Coach pensando..."):
            try:
                from groq import Groq
                client = Groq(api_key=st.secrets["GROQ_API_KEY"])
                response = client.chat.completions.create(
                    model="llama-3.3-70b-versatile",
                    messages=[{"role": "system", "content": user_context}] + st.session_state.chat_messages,
                    max_tokens=1024,
                )
                assistant_reply = response.choices[0].message.content
                st.session_state.chat_messages.append({"role": "assistant", "content": assistant_reply})
                st.rerun()
            except Exception as e:
                st.error(f"Error al conectar con AI Coach: {e}")

    st.markdown("<br>", unsafe_allow_html=True)
    with st.form(key="chat_form", clear_on_submit=True):
        col_input, col_btn, col_clear = st.columns([7, 1, 1])
        with col_input:
            user_input = st.text_input("", placeholder="Pregunta sobre tu entrenamiento, nutrición, rutinas...", label_visibility="collapsed")
        with col_btn:
            submitted = st.form_submit_button("Enviar", use_container_width=True)
        with col_clear:
            cleared = st.form_submit_button("🗑️ Limpiar", use_container_width=True)

    if submitted and user_input.strip():
        st.session_state.chat_messages.append({"role": "user", "content": user_input})
        st.rerun()
    if cleared:
        st.session_state.chat_messages = []
        st.rerun()

# Footer
st.markdown(f"""<div style="text-align:center;font-size:.72rem;color:#6b7a99;letter-spacing:1px;padding:1.5rem;">🔥 FireMuscle · {datetime.now().strftime("%H:%M:%S")} · {current_user}</div>""", unsafe_allow_html=True)
